'''
Ultimate final version... 11 May 2020..
Required packages
pip install openpyxl   For Excel control 3.0.3
pip install pywin32    For Mailing; pywin32   227; pywin32-ctypes      0.2.0
pip install pillow      FOR images  ;  7.1.2
pip install pyinstaller  For make EXE file or real software; PyInstaller         3.6
pyinstaller --name="MMCL ICX" --windowed --onefile hello.py
pyinstaller --name="MMCL ICX" -w  MMCL_monitoring_9_May_2020.py
pip install python-resize-image  # For resize images; python-resize-image 1.1.19
'''
import sqlite3
import console_out.console_out_gui_file
from console_out.console_out_gui_file import printCon
import ping_setting.ping_set,ping_setting.ping_setting_support,formula_gui
from formula_gui.formula_ import *
from ping_setting import *
from tkinter import messagebox
from os import getcwd
# from openpyxl import * # This code ambiguous for openpyxl_3.0.3  only for open() function.
from threading import Thread
import time,tkinter,subprocess,os,sys,threading,traceback
from tkinter import *
from datetime import datetime
from tkinter import Tk,filedialog,Frame,StringVar,Button,Label
from calendar import month_name
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tkinter.colorchooser import *
from functools import partial
import hashlib
import tkinter.ttk as ttk
from PIL import Image, ImageTk
from resizeimage import resizeimage

try:
    import win32com.client as win32
    from win32com.client import Dispatch
    from win32com.client import *
except Exception as e:
    exc_type, exc_value, exc_traceback = sys.exc_info();traceback.print_exception(exc_type, exc_value, exc_traceback)
    exc = traceback.format_exception(exc_type, exc_value, exc_traceback); [printCon(i, color='red') for i in exc]
    print(f"Exception Type:{e}; pywin32com module not installed")
    printCon(f"Exception Type:{e}; pywin32com module not installed",color='red')

###########################################################################################
# global conn,cursor
# conn = sqlite3.connect("all_data.db")
# cursor = conn.cursor()
# def myf():
#     while True:
#         print(f"Number of active thread: {threading.active_count()}")
#         print(f"Active thread  list: {threading.enumerate()}")
#         print(f"Current thread..... : {threading.current_thread()}")
#         currentThr = threading.current_thread()
#         print(f"Thread name.........: {currentThr.getName()}")
#         for tt in threading.enumerate():
#             print(tt.getName())
#         time.sleep(3)
# threading.Thread(target=myf, daemon=True).start()
###########################################################################################

class idd_class:

    def __init__(self):
        self.filename = "Open csv file "

    def upload(self):
        #self.filename = filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        filename = filedialog.askopenfilename(title = "Select IDD raw file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        printCon(f"IDD file dir: {filename}")
        if filename !="":
            self.filename = filename
            var_1.set(instance_idd_class.filename)

    def csv_file(self):
        try:
            fd1 = open(self.filename, "r");
        except Exception as e:
            print("Please in put file")
            printCon(f"Please in put file Exception type: {e}",color='red')
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            messagebox.showwarning("warning", "Invalid Input\nPlease in put first.44")
            return None

        if len(fd1.readline().split('","')) != 43:
            # text_var_out_folder.set("Wrong in put file.")
            messagebox.showwarning("warning", "Wrong input file.")
            return None

        fd1.seek(0, 0);
        length_fd1 = len(fd1.readlines());
        print('no of rows in csv file ',length_fd1)
        printCon(f"no of rows in csv file ,{length_fd1}")
        fd1.seek(0, 0);
        wb_idd = Workbook();
        sh_list = wb_idd.sheetnames;
        sh_list[0] ="idd raw";
        sh_list = wb_idd.sheetnames;
        ws_idd = wb_idd[sh_list[0]];
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [A]*6 +[B]*3+[C]+[D,D,E,I]+ [P]*3+[S]*7;  #to be delete column list
        printCon(f"Column to be deleted IDD2H: {del_col}")
        print("number of column = ",len(fd1.readline().split('","')))
        fd1.seek(0, 0);  # this command take file pointer at 0,0 position
        for itm in range(length_fd1):    # this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:         # delete unnecessary column from .csv file
                del row_list[item];
            row_list[-1] = row_list[-1][:-3]  # remove last list unnecessary ", sign which was created in csv file
            row_list.insert(0," ")  # insert first space for indent
            for item1 in range(len(row_list)):  # Convert string to integer
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                index_connect_number = row_list.index("Connect Number")
                index_attempt_number = row_list.index("Attempt Number")
                index_Answer_Number  = row_list.index("Answer Number")
                index_Answer_Time    = row_list.index("Answer Time")
                row_list.append("ASR")
                row_list.append("ACD")
                row_list.append("CCR")
            else:
                max_ro = itm;
                row_list.append(f"={chr(65 + index_Answer_Number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                row_list.append(f"={chr(65 + index_Answer_Time)}{max_ro + 1}/{chr(65 + index_Answer_Number)}{max_ro + 1}/60")
                row_list.append(f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")

            ws_idd.append(row_list);

        fd1.close();
        min_row = ws_idd.min_row
        min_col = ws_idd.min_column
        max_row = ws_idd.max_row
        max_col = ws_idd.max_column
        #col_letter = get_column_letter(max_col)

        ''''
        Setting style to the border and fonts, 
        '''
        cell_range = [min_row,min_col+1,max_row,max_col]  #set style all sheet.
        mmcl_domestic_ISD_report_class.set_border(None, ws=ws_idd, cell_range=cell_range)

        top_cell_range = [1,2,1,ws_idd.max_column]  # set style top row/ Header row
        mmcl_domestic_ISD_report_class.set_border(None, ws=ws_idd, cell_range=top_cell_range, bd_sty="medium", font_weight=True, bg_color=header_color)  # Border width = medium valid !!

        top_cell_range = [2,ws_idd.max_column-2,ws_idd.max_row,ws_idd.max_column]  # set style ASR,ACD,CCR
        mmcl_domestic_ISD_report_class.set_border(None, ws=ws_idd, cell_range=top_cell_range, bd_sty="thin", font_weight=True, bg_color=side_color)

        rd = ws_idd.row_dimensions[1]  # get dimension for row 3
        rd.height = 48  # value in points, there is no "auto"
        sd1 = [4,16,14,9,12,15,15,12,10,10,11,11,10,12,11,11,11, 13,11,11,9,9,9]
        print(len(sd1))
        for sd in range(1,ws_idd.max_column+1):
            cd = ws_idd.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd-1]
        ws_idd.freeze_panes = "A2"  # make freeze before B2
        # ws_idd.insert_cols(1)  # this will insert correctly but little problem with my fix excel formulas

        #  floating number will show up to two decimal point
        num_f = ws_idd.iter_rows(min_row=2,min_col=ws_idd.max_column-2,max_row=ws_idd.max_row,max_col=ws_idd.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        # calculate upto time
        for itm in range(2,30):
            a = int(ws_idd.cell(itm,4).value)    # 'D' column contain Hours informations
            b = int(ws_idd.cell(itm+1,4).value)     # 'D' column contain Hours informations
            if a>b:
                break
        if a<23:
            time_upto = datetime.strptime(f"{a+1}", "%H").strftime("%I %p")
        elif a==23:
            time_upto = '11 PM'

        # save file
        # set output folder
        if os.path.isdir(out_folder):
            pass
        else:
            Options.change_out_folder()

        # Change output folder
        global out_folder_dir
        try:
            os.chdir(out_folder_dir)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("output folder not set ",e)
            printCon(f"output folder not set {e}",color='red')
        date_file = (ws_idd["C5"].value).split("-")     # 'C' column contain date informations
        print(date_file)
        printCon(f"date_file = ,{date_file}")
        month_file = month_name[int(date_file[1])]
        name = f"IDD Report {date_file[2]} {month_file} {date_file[0]} (Every 2 Hours).xlsx"
        wb_idd.save(name)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        try:
            xl = Dispatch("Excel.Application")
            wb2 = xl.Workbooks.Open(Filename=path2)
            xl.Visible = False  # speed up process also
            wb2.Save()  # Save and over lap the original file
            wb2.Close(True)
            xl.Quit()
        except Exception:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
        # set currnt directory as previous
        os.chdir(current_working_dir)


        print("IDD Report Done.\n","*"*20,"\n")
        console_out.console_out_gui_file.set_in_console("IDD Report Done.\n")
        # set GUI bottom Label textvariable . var_3 is a tkinter.StringVar() object
        # Print completion message
        # var_3.set("IDD Report Done. Do Again")
        # lab_1["text"]="IDD Report Done." # textvariable and text will not work simultaneously
        text_var_out_folder.set("IDD Report Done.") # placed in parent window bottom Label

        # except Exception: print("File not found \n Or Wrong file selected \nOut file has opened")

        # mailing purpose only..
        idd_class.filename_mail_2h = name
        idd_class.filename_mail_2h_t = [date_file[0],month_file, date_file[2],f'{time_upto}']   # [ year, month, day , upto ]
        if auto_mail_enable.get() == 1:
            mail_instance.idd_every_two_hours()
        return None

class ccr_class:

    def __init__(self):
        self.countries = ["India", "Malaysia", "Saudi Arabia", "Singapore", "UAE"]
        self.all_countries = []  # To  store all country names
        self.filename = "Open csv file "

    def upload(self):
        filename = filedialog.askopenfilename(title = "Select file for CCR Check",filetypes = (("csv files","*.csv"),("all files","*.*")))
        #self.filename = filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        printCon(f"CCR file dir: {filename}")
        if filename !="":
            self.filename = filename
            var1.set(instance_ccr_class.filename)
        return None

    def csv_file(self):
        self.all_countries = []  # To  store all country names
        countries = ["India", "Malaysia", "Saudi Arabia", "Singapore", "UAE"]
        try:
            fd1 = open(self.filename, "r")
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("Please in put file")
            printCon(f"Please in put file Exception type: {e}",color='red')
            # text_var_out_folder.set("Please input first.")
            messagebox.showwarning("warning", "Invalid Input\nPlease in put first.")
            return None

        if len(fd1.readline().split('","')) != 46:
            # text_var_out_folder.set("Wrong in put file.")
            printCon(f"warning Wrong input file.",color='red')
            messagebox.showwarning("warning", "Wrong input file.")
            return None

        fd1.seek(0, 0);
        length_fd1 = len(fd1.readlines());
        fd1.seek(0, 0);
        wb_ccr = Workbook();
        # Dhaka = wb_in_kpi.create_sheet("Dhaka", 0)  # insert at first position
        sh_list = wb_ccr.sheetnames;
        ccr = wb_ccr[sh_list[0]];
        ccr.title = "CCR Check"
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [0]*6 +[10]*9 + [17]*14;
        print("CCR column Delete list : ",del_col);
        fd1.seek(0, 0); #this command take file pointer at 0,0 position
        abc = True
        for itm in range(length_fd1): #this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:  # delete unnecessary column from .csv file
                del row_list[item];
            for item1 in range(len(row_list)):  #convert str to int
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                # row_list[len(row_list) - 1] = "CCR";  # for the first row put CCR at the top ---------------------------------------------------------------------------
                row_list.append("CCR value")
                # print(f"'habib------',{row_list},{len(row_list)}")
                index_connect_number = row_list.index("Connect Number")  # find index of connect number to calculate ccr
                index_attempt_number = row_list.index("Attempt Number")
                index_Answer_Number = row_list.index("Answer Number")
                index_Answer_Time = row_list.index("Answer Time")
                index_CURRENT_TIME = row_list.index("CURRENT TIME")  # to filter time, this is index of a list not xlsx
            else:
                try:
                    row_list.append(int(row_list[index_connect_number])/int(row_list[index_attempt_number])*100) # new------------------------
                except ZeroDivisionError as b:
                    row_list.append(0)
                    if abc == True:
                        printCon(f"zero division error; Exception type:{b}",color='red')
                        abc = False
                except Exception as e:
                    exc_type, exc_value, exc_traceback = sys.exc_info()
                    traceback.print_exception(exc_type, exc_value, exc_traceback)
                    exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
                    [printCon(i, color='red') for i in exc]
                    print("Unknown Error ")

            # make a list of all countries
            if row_list[0] == "Object Description":pass
            elif row_list[0] in countries: pass
            elif row_list[0] not in self.all_countries:
                self.all_countries.append(row_list[0])


            if row_list[0] in self.countries:  # skip unnecessary countries
                max_ro = ccr.max_row
                row_list.append(f"  ")
                row_list.append(f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                row_list.append(f"={chr(65 + index_Answer_Number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                row_list.append(f"={chr(65 + index_Answer_Time)}{max_ro + 1}/{chr(65 + index_Answer_Number)}{max_ro + 1}/60")
                ccr.append(row_list)
            elif row_list[0] == "Object Description":
                row_list.append(" ") # -------------------------------------------------------------------------------------------------------------------------------
                row_list.append("CCR Equ") # -------------------------------------------------------------------------------------------------------------------------------
                row_list.append("ASR Equ") # -------------------------------------------------------------------------------------------------------------------------------
                row_list.append("ACD Equ") # -------------------------------------------------------------------------------------------------------------------------------
                ccr.append(row_list)

        fd1.close()

        printCon(f"Selected contries : {self.countries}")

        self.all_countries.sort()
        self.all_countries = countries + self.all_countries # place important countries first then other countries sort form.
        all_countries_file = open('all_countries.txt','w+')
        for i in range(len(self.all_countries)):
            all_countries_file.write(f"{self.all_countries[i]} \n")

        # all_countries_file.write(str(all_countries))
        all_countries_file.close()

        ccr.auto_filter.ref = f"A1:Z{ccr.max_column}"
        # ccr.auto_filter.add_filter_column(0, ["India", "Malaysia", "Saudi Arabia","Singapore","UAE"])
        ccr.auto_filter.add_filter_column(0, self.countries)
        time_slot = []
        time_slot_2 = []
        time_slot_range = []
        for fil_1 in range(2,30):
            var23 = ccr.cell(fil_1, index_CURRENT_TIME+1).value # since list index start with 0,
            # and xlsx index start with 1 that's why here add 1 (one).
            time_slot_2.append(var23)
        print("time slot_2 = ",time_slot_2)
        for fil_3 in range(27):
            time_slot.append(int(time_slot_2[fil_3][:2]))
        print("time slot = ", time_slot)
        for fil_2 in range(24):
            if time_slot[fil_2]>time_slot[fil_2+1]:
                time_slot_range.append(time_slot_2[fil_2-2])
                time_slot_range.append(time_slot_2[fil_2-1])
                time_slot_range.append(time_slot_2[fil_2])
                break;
        print(f"Last three hours:  {time_slot_range}")
        printCon(f"Last three hours:  {time_slot_range}")
        ccr.auto_filter.add_filter_column(index_CURRENT_TIME, time_slot_range)

        # set border and cell property to Header cells
        mmcl_domestic_ISD_report_class.set_border(None, ws=ccr, cell_range=[1, 1, 1, ccr.max_column], bd_sty="medium", font_weight=False, bg_color=header_color)
        # set border to data cells
        mmcl_domestic_ISD_report_class.set_border(None, ws=ccr, cell_range=[2, 1, ccr.max_row, ccr.max_column])
        # set font color Red and weight bold it CCR less than 80%
        low_ccr_list = []
        for color1 in ccr.iter_rows(min_row=3, min_col=1):  # red mark ccr less than 80%
            if int(color1[ccr.max_column - 5].value) < 70:
                for color2 in color1:
                    color2.font = Font(bold=False, color="ff0000")
                    low_ccr_list.append(color2.value)
        print(f"Low CCR count today = {len(low_ccr_list)/ccr.max_column}")
        console_out.console_out_gui_file.set_in_console(f"Low CCR count today = {len(low_ccr_list)/ccr.max_column}")

        sd1 = [16, 14, 25, 16, 12, 14, 11, 11, 10, 12, 13, 12, 12, 11, 13, 11, 13, 11, 11, 9, 9, 9]
        for sd in range(1, ccr.max_column + 1):
            cd = ccr.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        rd = ccr.row_dimensions[1]  # get dimension for row 3
        rd.height = 45  # value in points, there is no "auto"
        ccr.freeze_panes = "E2"  # make freeze before E2

        #  floating number will show up to two decimal point
        num_f = ccr.iter_rows(min_row=2,min_col=ccr.max_column-1,max_row=ccr.max_row,max_col=ccr.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        # Save file
        # set output folder
        if os.path.isdir(out_folder):
            pass
        else:
            Options.change_out_folder()

        # Change output folder
        global out_folder_dir
        try:
            os.chdir(out_folder_dir)
            console_out.console_out_gui_file.set_in_console(f"output folder changed... to {out_folder}")
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"output folder not set , {e}")
            printCon(f"output folder not set , {e}", color='red')

        ccr_out_file = f"CCR_ISD_Check up to {time_slot_range[2]}.xlsx"
        ccr_out_file = ccr_out_file.replace(":","_")  # file name do not contain : sign
        print(f"Out file name ccr : {ccr_out_file}")
        wb_ccr.save(ccr_out_file)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        if intVar_ccr_auto_open.get() == 1:
            path1 = os.getcwd() + os.sep
            path2 = os.path.join(path1, ccr_out_file)
            try:
                xl = Dispatch("Excel.Application")
                wb2 = xl.Workbooks.Open(Filename=path2)
                xl.Visible = True  # speed up process also
                wb2.Save()  # Save and over lap the original file
                console_out.console_out_gui_file.set_in_console(f"CCR file opening in Excel....")
                # wb2.Close(True)
                # xl.Quit()
            except Exception as ex:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                traceback.print_exception(exc_type, exc_value, exc_traceback)
                exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
                [printCon(i, color='red') for i in exc]
                printCon(f"MS Execl may be not installed in this machine: Exception type: {ex}", color='red')
        else:pass

        # Set output folder as previous
        os.chdir(current_working_dir)

        # set GUI bottom Label textvariable . var_3 is a tkinter.StringVar() object
        # Print completion message
        # lab_1["text"] = "CCR Check Done"
        text_var_out_folder.set("CCR Check Done")  # placed in parent window bottom Label
        print(f"CCR Check file has Done.\n" + "*"*20 + "\n")
        printCon(f"CCR Check file has Done.\n" + "*"*20 + "\n")
        return None

    def country_selection(self):
        # print(self.countries)
        print(f"self.all_countries 1 = {self.all_countries}")
        if self.all_countries.__len__() == 0:
            try:
                file = open("all_countries.txt","r+")
            except Exception:
                messagebox.showwarning("warning", "No Previous data exist", parent=top)
                return None
            for item in file.readlines():
                self.all_countries.append(item[:-2])
            print(f"self.all_countries 2 = {self.all_countries}")
            file.close()

        # Gui started here
        select_countries_gui = Toplevel()
        select_countries_gui.title("Select Countries")
        select_countries_gui.resizable(width=False, height=False)
        Lb1 = Listbox(select_countries_gui, selectmode=MULTIPLE, width=30,height=20,font="Times")


        def change():
            list1 = []
            for i in list(Lb1.curselection()):
                list1.append(Lb1.get(i))
            print(list1)
            if len(list1) != 0:
                self.countries = list1
            select_countries_gui.destroy()


        for item, itm in zip(self.all_countries,range(len(self.all_countries))):
            Lb1.insert(itm, item)

        fram = Frame(select_countries_gui)
        btn = Button(fram, text="Select", command=change,width=15,font="Times")
        btn2 = Button(fram, text="Clear All", command=lambda :Lb1.selection_clear(0,"end"),width=15,font="Times")
        btn4 = Button(fram, text="Quit", command=lambda : select_countries_gui.destroy(),width=15,font="Times")

        fram.grid(column=2,row=2)

        Lb1.grid(column=2,row=1)
        btn.grid(column=2,row=2)
        btn2.grid(column=3,row=2)
        btn4.grid(column=5,row=2)

        for itm in range(len(self.countries)):
            Lb1.select_set(self.all_countries.index(self.countries[itm]))
        select_countries_gui.mainloop()

class mmcl_domestic_ISD_report_class:

    # make class variable
    db_var_summary_list = []
    wb_kpi_raw_in = Workbook()  # not used yet
    wb_kpi_raw_out = Workbook()
    filename = ' Open File .. .. .. .'
    all_city_trunk = [];
    ver_sum = 0;
    def __init__(self):
        self.wb_raw_data = Workbook()
        self.ws_raw_data = self.wb_raw_data[self.wb_raw_data.sheetnames[0]]
        self.wb_report = Workbook();
        temp_rem = self.wb_report.active
        self.wb_report.remove(temp_rem)
        mmcl_domestic_ISD_report_class.index_of_answer_time = 0 # this is a class variable. same memory point for all functions
        self.total_time = 0

    def upload_kpi(self,kpi_top):
        filename = filedialog.askopenfilename(title = "Select incoming KPI raw CSV file",filetypes =
        (("csv files","*.csv"),("all files","*.*")),parent=kpi_top)
        if filename != '':
            self.filename = filename
            raw_in_kpi_file_path.set(raw_kpi_in_instance.filename)
            #raw_kpi_in_instance.uploaded_cvs_to_xlsx_kpi()
        printCon(f"incoming KPI raw CSV file: {filename}")
        return None

    def upload_kpi_out(self,kpi_top):
        filename = filedialog.askopenfilename(title = "Select outgoing KPI Raw CSV File",filetypes =
        (("csv files","*.csv"),("all files","*.*")),parent=kpi_top)
        if filename != '':
            self.filename = filename
            raw_out_kpi_file_path.set(raw_kpi_out_instance.filename)
            # raw_kpi_out_instance.uploaded_cvs_to_xlsx_kpi()
        printCon(f"Outgoing KPI raw CSV file: {filename}")
        return self.filename

    def upload_idd(self,kpi_top):
        filename = filedialog.askopenfilename(title = "Select idd Raw CSV File",filetypes = (("csv files","*.csv"),
                                                    ("all files","*.*")),parent=kpi_top)
        if filename != '':
            self.filename = filename
            raw_idd_file_path.set(self.filename)
        printCon(f"IDD raw CSV file: {filename}")
        return self.filename

    def uploaded_cvs_to_xlsx_kpi(self):
        global kpi_top
        try:
            fd1 = open(self.filename, "r");
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("Please in put file")
            messagebox.showwarning("warning", "Invalid Input\nPlease input first.", parent=kpi_top)
            return None

        if len(fd1.readline().split('","')) != 45:
            messagebox.showwarning("warning", "Wrong input file.", parent=kpi_top)
            return None

        fd1.seek(0, 0);
        length_fd1 = len(fd1.readlines());
        fd1.seek(0, 0);
        wb_kpi = Workbook();
        sh_list = wb_kpi.sheetnames;
        ws_kpi = wb_kpi[sh_list[0]]
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [A]*6 +[B]*3+[C,C]+[F,G,K]+ [R]*3+[U,U]+[V]*4;  #to be delete column list
        fd1.seek(0, 0);  # this command take file pointer at 0,0 position
        for itm in range(length_fd1):    # this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:         # delete unnecessary column from .csv file
                del row_list[item];
            row_list[-1] = row_list[-1][:-3]  # remove last list unnecessary ", sign which was created in csv file
            # row_list.insert(0," ")  # insert first space for indent
            for item1 in range(len(row_list)):  # Convert string to integer
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                index_Answer_Time    = row_list.index("Answer Time")
                mmcl_domestic_ISD_report_class.index_of_answer_time = index_Answer_Time   # this is a class variable. same memory point for all functions

            ws_kpi.append(row_list);
        fd1.close()
        self.wb_raw_data = wb_kpi;
        self.ws_raw_data = wb_kpi[wb_kpi.sheetnames[0]]
        # wb_kpi.save("_in_kpi_cut file.xlsx")  # save data only for testing purpose
        # os.chdir(os.getcwd()+ "\ new")
        return ws_kpi

    def trunk_finder(self,raw_kpi_ws):
        list1 = []  # Local variable
        list2 = []
        for itm in range(1, raw_kpi_ws.max_row):
            if raw_kpi_ws.cell(itm, 1).value not in list1:
                list1.append(raw_kpi_ws.cell(itm, 1).value)
                list2.append(itm)
        printCon(f"Trunk list: {dict(zip(list1, list2))}")
        return list1, dict(zip(list1, list2))

    def city_trunk_finder(self,list_kpi_trunk, key):
        list_kpi_trunk_city = []
        for itm in list_kpi_trunk:
            if not key != itm[-2:]:  # for Dhaka key = DH; khulna key=KH;
                sum = self.answer_time_sum(itm)
                if sum>0:
                    list_kpi_trunk_city.append(itm)
        # print("list_kpi_trunk_city ",list_kpi_trunk_city)
        printCon(f"Valid Trunk sort by Cities: {list_kpi_trunk_city}")
        return list_kpi_trunk_city

    def city_report(self,city,list_kpi_trunk_city,ws_raw_data,dic_trunk_col_raw,time_range):
        # Creation of city Worksheet instance
        temp_summary_db = 0
        if f"{city}" in self.wb_report.sheetnames:
            temp_ws = self.wb_report[f"{city}"]
            self.wb_report.remove(temp_ws)  # remove sheet if it already exist
            self.wb_report.create_sheet(f"{city}")  # a new worksheet will creates with city name title
        else:
            self.wb_report.create_sheet(f"{city}")  # a new worksheet will creates with city name title
        City = self.wb_report[f"{city}"]  # a new worksheet will be in City variable.
        # City.title(f"{city}") # value less code. title already given when created.!!

        # First row of the city report: Header of city report.
        header_row = [ws_raw_data.cell(1,hr).value for hr in range(1,self.ws_raw_data.max_column+1)] # list com
        header_row.insert(0," ")
        header_row.append("ASR")
        header_row.append("ACD")
        header_row.append("CCR")
        header_row[3] = "Start Time"    # "Start Time" is more meaningful than "Last Time"
        header_row[4] = "End Time"      # "End Time" is more meaningful than "End Time"
        for itm,itm1 in zip(header_row,range(1,1+len(header_row))): City.cell(1,itm1).value = itm
        #City.append(header_row)
        cell_range_h = [1,2,1,len(header_row)]
        self.set_border(City, cell_range=cell_range_h,bd_sty="medium",font_weight=True,bg_color=header_color)

        index_connect_number = header_row.index("Connect Number")
        index_attempt_number = header_row.index("Attempt Number")
        index_Answer_Number = header_row.index("Answer Number")
        index_Answer_Time = header_row.index("Answer Time")
        index_date = header_row.index("DATE")

        # self.wb_report.save("test_report.xlsx")

        print(dic_trunk_col_raw)
        printCon(dic_trunk_col_raw)
        length = len(list_kpi_trunk_city)
        # print("length of city list ******** ", length,list(range(length)))
        print(f"city list {list_kpi_trunk_city}")
        printCon(f"city list : {list_kpi_trunk_city}")

        list1 = {}

        # for itm, item in list_kpi_trunk_city,range(length):  # nor working
        for itm, item in zip(list_kpi_trunk_city,range(length)):  # for all trunk in a city.
            # range used only for numaric variable to use. useless !!
            start_row = dic_trunk_col_raw.get(itm)  # this will give the starting point of the raw data
            for itm1 in range(1,time_range+1):  # this for: complete a single trunk report
                for itm2 in range(1,ws_raw_data.max_column+1):  # this for: complete a single row of a report
                    row = start_row + (itm1-1)
                    col = itm2
                    temp1 = ws_raw_data.cell(row, col).value;  # collect value form raw file.
                    row2 = time_range*item + (itm1-1)+2*(item+1)  #
                    col2 = itm2 +1 # plus 1 for report start with second column
                    City.cell(row2, col2).value = temp1;  # put value in report file
                    if col2 == 12: temp_summary_db += temp1 # only for make database of answer number

                max_ro = itm1;
                row_number = max_ro + 1+(item * (time_range + 2))
                City.cell(row2, col2 + 1).value = f"={chr(65 + index_Answer_Number)}{row_number}/{chr(65 + index_attempt_number)}{row_number}*100"
                City.cell(row2, col2 + 2).value = f"={chr(65 + index_Answer_Time)}{row_number}/{chr(65 + index_Answer_Number)}{row_number}/60"
                City.cell(row2, col2 + 3).value = f"={chr(65 + index_connect_number)}{row_number}/{chr(65 + index_attempt_number)}{row_number}*100"
            list1[f"{itm}"] = f"L{row2-time_range+1}:L{row2}"  # plus +1 means -> from 10 to 20 = 11
            City.merge_cells(start_row=row2-time_range+1, start_column=2, end_row=row2,end_column=2)
            City.merge_cells(start_row=row2-time_range+1, start_column=3, end_row=row2,end_column=3)

            # apply style to the sheets
            cell_range = [row2-time_range+1,2,row2,col2]   # apply to all values
            self.set_border(City,cell_range=cell_range,wraptext=False)  # only this function need 2.5 seconds to set border !!

            # merged cell must be bold. . .
            cell_range = [row2 - time_range + 1, 2, row2, 3]
            self.set_border(City, cell_range=cell_range, font_weight=True)

            # set side/ACD,CCR cell styles
            cell_range_cal = [row2-time_range+1,col2 + 1,row2,col2 + 3]    # apply to ASR, ASD, CCR columns
            self.set_border(City,cell_range=cell_range_cal,bd_sty="medium",bg_color=side_color,wraptext=False)

        print(f"temp_summary_db . . {temp_summary_db}")
        print(f"index_Answer_Number = {index_Answer_Number}")
        self.db_var_summary_list.append(temp_summary_db)
        # set Column width
        sd1 = [1, 29, 12, 8, 8, 9, 11, 11, 12, 12, 9, 11, 9, 9, 9, 9, 9, 9, 10, 12, 10, 10,11,8,8,8,10]
        printCon(f"Set width of Excel file columns Domestic KPI: {sd1}")
        # print(len(sd1))
        for sd in range(1, City.max_column + 1):
            cd = City.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        rd = City.row_dimensions[1]  # get dimension for row 1
        rd.height = 45  # value in points, there is no "auto"
        City.freeze_panes = "A2"  # make freeze before A2, means up to A1,B1,C1 . . .. .

        #  floating number will show up to two decimal point
        num_f = City.iter_rows(min_row=2, min_col=City.max_column - 2, max_row=City.max_row, max_col=City.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '###0.00'

        return list1

    def make_kpi_report(self):
        self.db_var_summary_list.clear() # except this code will show error when second time run this software... previous data need to be clear
        var_stop = self.uploaded_cvs_to_xlsx_kpi()
        if var_stop == None:return None # stop program if invalid input occurs
        # renew self.wb_report at first, if accidently click double on make report
        self.wb_report = Workbook();
        temp_rem = self.wb_report.active
        self.wb_report.remove(temp_rem) # remove the atumatic created first Sheet

        self.total_time = 0 # Total domestic Answer time set 0
        time1=time.time()
        kpi_path = raw_kpi_in_instance.filename;
        # print(kpi_path)
        sheet_names = self.wb_raw_data.sheetnames;

        raw_kpi_ws = self.wb_raw_data[sheet_names[0]]
        # print(raw_kpi_ws.max_row)
        list_kpi_trunk, self.dic_trunk_col_raw = self.trunk_finder(raw_kpi_ws)  # trunk and row in raw file.
        # print(self.dic_trunk_col_raw)
        printCon(f"Dictionary of trunk and rows: {self.dic_trunk_col_raw}")
        # exit()
        time_range = self.time_range_cal(raw_kpi_in_instance.ws_raw_data)  # time range means up to XX:XX PM

        # list_object_description = self.object_description_finder(raw_kpi_ws)
        list_city_kpi_trunk=[]      # only trunks which have Answer Time value of voice call.
        list_city_kpi_trunk.append(self.city_trunk_finder(list_kpi_trunk, "DH"))
        list_city_kpi_trunk.append(self.city_trunk_finder(list_kpi_trunk, "CH"))
        list_city_kpi_trunk.append(self.city_trunk_finder(list_kpi_trunk, "KH"))
        self.all_city_trunk = [list_city_kpi_trunk[0],list_city_kpi_trunk[1],list_city_kpi_trunk[2]]
        # insert Value for Dhaka
        # self.kpi_report = Workbook();

        dic_dh = self.city_report(city="Dhaka",list_kpi_trunk_city=list_city_kpi_trunk[0],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range=time_range)
        dic_ch = self.city_report(city="CTG",list_kpi_trunk_city=list_city_kpi_trunk[1],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range=time_range)
        dic_kh = self.city_report(city="Khulna",list_kpi_trunk_city=list_city_kpi_trunk[2],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range=time_range)


        all_dic = [dic_dh,dic_ch,dic_kh]
        self.summary(all_dic)

        # this block of code only for make database
        # self.db_var_summary_list.append(self.ws_raw_data.cell(4,2).value)
        self.db_var_summary_list.insert(0,self.ws_raw_data.cell(4,2).value)
        print(f"db_var_summary_list . . {self.db_var_summary_list}")
        printCon(f"db_var_summary_list . . {self.db_var_summary_list}")
        if self.time_range == 24:
            self.make_database_kpi_summary() # to send summary data to the database excel file
        self.db_var_summary_list.clear()

        # save file with name
        # set output folder
        if os.path.isdir(out_folder):
            pass
        else:
            Options.change_out_folder()

        # Change output folder
        global out_folder_dir
        try:
            os.chdir(out_folder_dir)
            printCon(f"out put folder: {out_folder_dir}")
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("output folder not set ", e)

        x = [0,0]
        for itm in range(2,30):
            x[0] = int(self.ws_raw_data.cell(itm,3).value.split(":")[0])
            x[1] = int(self.ws_raw_data.cell(itm+1,3).value.split(":")[0])
            if x[0]>x[1]:
                break
        # second row and second column has date.
        date_file = self.ws_raw_data.cell(2, 2).value.split("-")
        month_file = month_name[int(date_file[1])]
        if x[0] < 23:
            time_upto = datetime.strptime(f"{x[0]+1}", "%H").strftime("%I %p")
            name = f"ICX_KPI Report {date_file[2]} {month_file} {date_file[0]} (Upto {time_upto}).xlsx"
            time_upto = f" up to {time_upto}"
        else:
            name = f"ICX_KPI Report {date_file[2]} {month_file} {date_file[0]}.xlsx"
            time_upto = ''
        self.wb_report.save(name)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        try:
            xl = Dispatch("Excel.Application")
            wb2 = xl.Workbooks.Open(Filename=path2)
            xl.Visible = False  # speed up process also
            wb2.Save()  # Save and over lap the original file
            wb2.Close(True)
            xl.Quit()
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"Exception occurs...: {e} \n MS Excel may be not installed in this PC..")
            printCon(f"Exception occurs...: {e} \n MS Excel may be not installed in this PC..", color='red')
        # set current dir as previous
        os.chdir(current_working_dir)

        # Mailing purpose
        mmcl_domestic_ISD_report_class.nam_temp_kpi = name
        mmcl_domestic_ISD_report_class.nam_temp_kpi_t = [date_file[0], month_file, date_file[2], time_upto]
        if auto_mail_enable.get() == 1:
            mail_instance.kpi()

        print(f"Total domestic Answer time = , {self.total_time/60}")
        printCon(f"Total domestic Answer time =  {self.total_time/60}")
        printCon("file save successfully !! ")
        print("file save successfully !! ")

        time2=time.time()
        print("KPI report done, time = ",time2-time1)
        kip_conform.set(f" KPI report done \nTime needed: {time2-time1}  Total Minutes: {self.total_time/60}")
        return None

    def make_IOS_ISD_report(self):
        self.db_var_summary_list.clear()
        global dont_mail;dont_mail=1
        var_stop = raw_kpi_out_instance.uploaded_cvs_to_xlsx_kpi()
        if var_stop == None:return None # stop program if invalid input occurs
        if raw_kpi_in_instance.ws_raw_data['A1'].value == None:     # to make IOS ISD report must make KPI report first.
            messagebox.showwarning("warning", "Make KPI report first.", parent=kpi_top)
            return None
        print(f"value of A1 raw kpi = {raw_kpi_in_instance.ws_raw_data['A1'].value}")
        printCon(f"value of A1 raw kpi = {raw_kpi_in_instance.ws_raw_data['A1'].value}")

        # renew self.wb_report at first, if accidently click double on make report
        self.wb_report = Workbook();
        temp_rem = self.wb_report.active
        self.wb_report.remove(temp_rem)     # remove the atumatic created first Sheet

        self.total_time = 0  # Total domestic Answer time set 0
        time1 = time.time()
        sheet_names = self.wb_raw_data.sheetnames;

        raw_kpi_out_ws = self.wb_raw_data[sheet_names[0]]
        # print(raw_kpi_out_ws.max_row)
        list_kpi_trunk, self.dic_trunk_col_raw = self.trunk_finder(raw_kpi_out_ws)  # trunk and row in raw file.
        # print(self.dic_trunk_col_raw)

        timeR = self.time_range_cal(raw_kpi_out_ws)  # time range means up to XX:XX PM
        timeR1 = self.time_range_cal(raw_kpi_in_instance.ws_raw_data)  # time range means up to XX:XX PM

        # report work as IOS report.
        IOS_trunk_list = [["1001-Roots_IOS_IN"],["1005-NovoTel_IOS_IN"],["1009-Btrac_IOS_IN"],["1013-MirTelecom_IOS_IN"],["1017-Global_Voice_IN"],["1021-Unique_IOS"],["1025-Digicon_IOS"]]
        IOS_list = ["Roots","NovoTel","Btrac","MirTelecom","GlobalVoice","Unique","Digicon"]
        # IOS_trunk_list.reverse(); IOS_list.reverse()
        printCon(f"IOS Trunk list: {IOS_trunk_list}")
        printCon(f"IOS list: {IOS_list}")
        dic_ios = [] # list of trunk and value of excel ranges
        for itm in range(len(IOS_list)):
            dic_ios.append(self.city_report(city=f"{IOS_list[itm]}", list_kpi_trunk_city=IOS_trunk_list[itm],ws_raw_data=raw_kpi_in_instance.ws_raw_data,dic_trunk_col_raw=raw_kpi_in_instance.dic_trunk_col_raw,time_range = timeR1))
        # print("dic_iso = ",dic_ios)

        #   Summation of IOS ISD calls by function
        self.summary_ios(dic_ios)

        #   ISD report
        ans_isd_trunk = [["1050-TeleTalk_ISD"],["1070-GP1_ISD"],["1080-Robi_ISD"],["1090-Banglalink_ISD"]]
        ans_isd_list = ["TaleTalk ISD","GP ISD","ROBI ISD","Banglalink ISD"]
        # ans_isd_list.reverse();  ans_isd_trunk.reverse();
        for itm in range(len(ans_isd_trunk)):
            self.city_report(city=f"{ans_isd_list[itm]}", list_kpi_trunk_city=ans_isd_trunk[itm],ws_raw_data=self.ws_raw_data,dic_trunk_col_raw=self.dic_trunk_col_raw,time_range = timeR)


        # verification test. kpi and ios file wrong input
        IOS_trunk_verification = ["1001-Roots_IOS_IN","1005-NovoTel_IOS_IN","1009-Btrac_IOS_IN","1013-MirTelecom_IOS_IN","1017-Global_Voice_IN","1021-Unique_IOS","1025-Digicon_IOS"]
        sum2=0
        for trunk in IOS_trunk_verification:
            row = self.dic_trunk_col_raw.get(trunk)
            ws = raw_kpi_in_instance.wb_raw_data[raw_kpi_in_instance.wb_raw_data.sheetnames[0]]
            a = mmcl_domestic_ISD_report_class.index_of_answer_time + 1
            b = self.time_range_cal(raw_kpi_in_instance.ws_raw_data)
            sum1 = 0
            for itm in range(b):
                sum1 += ws.cell(row + itm, a).value
            sum2 += sum1
        print(sum2/60,"new sum verification")

        # save file with name
        # set output folder
        if os.path.isdir(out_folder):
            pass
        else:
            Options.change_out_folder()

        # Change output folder
        global out_folder_dir
        try:
            os.chdir(out_folder_dir)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]

        x = [0, 0]
        for itm in range(2, 30):
            x[0] = int(self.ws_raw_data.cell(itm, 3).value.split(":")[0])
            x[1] = int(self.ws_raw_data.cell(itm + 1, 3).value.split(":")[0])
            if x[0] > x[1]:
                break
        # second row and second column has date.
        date_file = self.ws_raw_data.cell(2, 2).value.split("-")
        month_file = month_name[int(date_file[1])]
        if x[0] < 23:
            time_upto = datetime.strptime(f"{x[0] + 1}", "%H").strftime("%I %p")
            name = f"IOS & ISD Report {date_file[2]} {month_file} {date_file[0]} (Upto {time_upto}).xlsx"
            time_upto = f" up to {time_upto}"
        else:
            name = f"IOS & ISD Report {date_file[2]} {month_file} {date_file[0]}.xlsx"
            time_upto = ''
        self.wb_report.save(name)
        print("Total domestic Answer time = ", self.total_time / 60)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        try:
            xl = Dispatch("Excel.Application")
            wb2 = xl.Workbooks.Open(Filename=path2)
            xl.Visible = False  # speed up process also
            wb2.Save()  # Save and over lap the original file
            wb2.Close(True)
            xl.Quit()
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"Exception type: {e}; in make_IOS_report() pywin32 related issue")
            printCon(f"Excel not installed..Exception type: {e}; in make_IOS_report() pywin32 related issue", color='red')
        # set current dir as previous
        os.chdir(current_working_dir)

        time2 = time.time()
        print("KPI report done, time = ", time2 - time1)

        # Calculate total Answer time for outgoing kpi file
        ans_isd_trunk1 = ["1050-TeleTalk_ISD","1070-GP1_ISD","1080-Robi_ISD","1090-Banglalink_ISD"]
        for itm in ans_isd_trunk1: self.answer_time_sum(itm)

        print("self.total_time = ",self.total_time/60)
        print("sum2 = ",sum2/60)

        if abs(int(self.total_time/60) - int(sum2/60)) <= 15:   # if two sum difference is less than 15 minutes than its acceptable.
            kip_conform.set(f" IOS IDD report done \n Time needed: {time2-time1} Total Minutes : {self.total_time/60} (outgoing kpi)")
        else:
            kip_conform.set(f"Wrong Input file")
            messagebox.showerror("error", f"Error hh \n Wrong Input file value difference is {abs(int(self.total_time/60) - int(sum2/60))}mt",parent=kpi_top)
            dont_mail=0

        # for Mailing purpose
        mmcl_domestic_ISD_report_class.nam_temp_ios = name
        mmcl_domestic_ISD_report_class.nam_temp_ios_t = [date_file[0], month_file, date_file[2], time_upto]
        if auto_mail_enable.get() == 1 and dont_mail==1:
            mail_instance.ios_idd()

        self.db_var_summary_list.insert(0,self.ws_raw_data.cell(4,2).value)
        self.make_database_ios_summary()

        return None

    def answer_time_sum(self,trunk):
        row = self.dic_trunk_col_raw.get(trunk)
        ws = self.wb_raw_data[self.wb_raw_data.sheetnames[0]]
        a = mmcl_domestic_ISD_report_class.index_of_answer_time + 1
        b=self.time_range_cal(self.ws_raw_data)
        sum1 = 0
        for itm in range(b):
            sum1 += ws.cell(row+itm,a).value
        # print("summission of a trunk: ",trunk,sum1)
        self.total_time += sum1
        return sum1

    def time_range_cal(self,ws_raw_data):
        last_time = 3 # just random value
        for itm1 in range(1,10):
            if ws_raw_data.cell(1, itm1).value == "LAST TIME":
                last_time = itm1
                break
        list1 = []
        list2 = []
        for itm2 in range(2,40):
            list1.append(ws_raw_data.cell(itm2,last_time).value)
        for itm in list1:
            list2.append(int(itm[:2]))
        for itm3 in range(30):
            if list2[itm3]>list2[itm3+1]:
                time_range = list2[itm3]+1
                break
        self.time_range = time_range
        return time_range

    def summary(self, all_dic):
        global db_var_summary_list
        # remove sheet if exist already
        if "Summary" not in self.wb_report.sheetnames:
            self.wb_report.create_sheet(title="Summary",index=3)
        else:
            temp_ws = self.wb_report["Summary"]
            self.wb_report.remove(temp_ws)
            self.wb_report.create_sheet(title="Summary",index=3)
        ws_summary = self.wb_report["Summary"]

        #Border Style set
        bd_sty = "medium"
        font_weight = True

        # all_dic = list of dictionary of city trunks with value "Answer Time" ranges it will help to make sum.
        list1 =["Zone","Trunk ID","Total Time (Sec)","Total Time (Min)"]
        list2 = [["Zone","Total Domestic Time (Min)"],["Dhaka",0],["CTG",0],["Khulna",0]]

        # self.ws_summary.cell(2,7).value = list1

        # variables that used
        col_ind = 8     # How many column will be empty before summary calculation
        len0 = 3        # position of Dhaka trunk summary
        len1 = len(self.all_city_trunk[0]) + len0 + 4   # CTG trunk summary starts
        len2 = len(self.all_city_trunk[1]) + len1 + 4
        len3 = len(self.all_city_trunk[2]) + len2 + 4
        length = [len0,len1,len2,len3]
        #make class variable
        mmcl_domestic_ISD_report_class.check_loop = 0

        # =SUM(Dhaka!L2: L25)
        # for itm,itm2 in zip(list1,range(len(list1))):   # print header of each summary
        for itm2,itm in enumerate(list1):   # print header of each summary
            ws_summary.cell(len0, col_ind+itm2).value = itm  # print header of each summary
            ws_summary.cell(len1, col_ind+itm2).value = itm
            ws_summary.cell(len2, col_ind+itm2).value = itm
        # Dhaka  # Starting column -3 means first col of summary header
        cell_range_dhH = [len0,(col_ind+itm2)-3,len0,col_ind+itm2]
        # CTG
        cell_range_chH = [len1, (col_ind + itm2) - 3, len1, col_ind + itm2]
        # Khulna
        cell_range_khH = [len2, (col_ind + itm2) - 3, len2, col_ind + itm2]

        # ******************************************************************************************
        for var1, var2 in zip(self.all_city_trunk[0], range(len(self.all_city_trunk[0]))):
            mmcl_domestic_ISD_report_class.check_loop += 1
            list_dh = ["Dhaka", var1, f"=SUM(Dhaka!{all_dic[0].get(var1)})", f"=SUM(Dhaka!{all_dic[0].get(var1)})/60"]
            for var3, var4 in zip(list_dh, range(4)):
                ws_summary.cell((len0+1) + var2, var4 + col_ind).value = var3
        # ws_summary.merge_cells("B1:B5") # this syntax also valid.
        ws_summary.merge_cells(start_row=(len0+1), start_column=col_ind, end_row=(len0+1) + var2, end_column=col_ind)
        sum_col = get_column_letter(var4 + col_ind)
        ws_summary.cell((len0+1) + var2+1,var4 + col_ind).value = f"=SUM({sum_col}{(len0+1)}:{sum_col}{(len0+1) + var2})"  # make summary of same city trunks in Minutes
        list2[1][1] = f"={sum_col}{(len0+1) + var2+1}"
        # cell ranges for simple style set
        cell_range_dh1 = [(len0+1), col_ind, (len0+1) + var2, var4 + col_ind]
        # so sad this function calls for only one cell !!
        cell_range_dh2 = [(len0 + 1)+var2+1, var4 + col_ind, (len0 + 1) + var2+1, var4 + col_ind]

        for var1, var2 in zip(self.all_city_trunk[1], range(len(self.all_city_trunk[1]))):
            list_dh = ["CTG", var1, f"=SUM(CTG!{all_dic[1].get(var1)})",
                       f"=SUM(CTG!{all_dic[1].get(var1)})/60"]
            for var3, var4 in zip(list_dh, range(4)):
                ws_summary.cell((len1 + 1) + var2, var4 + col_ind).value = var3
        ws_summary.merge_cells(start_row=(len1 + 1), start_column=col_ind, end_row=(len1 + 1) + var2, end_column=col_ind)
        ws_summary.cell((len1+1) + var2+1,var4 + col_ind).value = f"=SUM({sum_col}{(len1+1)}:{sum_col}{(len1+1) + var2})"  # make summary of same city trunks in Minutes
        list2[2][1] = f"={sum_col}{(len1+1) + var2+1}"
        cell_range_ch1 = [(len1 + 1), col_ind, (len1 + 1) + var2, var4 + col_ind]
        cell_range_ch2 = [(len1 + 1) + var2+1,  var4 + col_ind, (len1 + 1) + var2+1 , var4 + col_ind]

        for var1, var2 in zip(self.all_city_trunk[2], range(len(self.all_city_trunk[2]))):
            list_dh = ["Khulna", var1, f"=SUM(Khulna!{all_dic[2].get(var1)})",
                       f"=SUM(Khulna!{all_dic[2].get(var1)})/60"]
            for var3, var4 in zip(list_dh, range(4)):
                ws_summary.cell((len2 + 1) + var2, var4 + col_ind).value = var3
        ws_summary.merge_cells(start_row=(len2 + 1), start_column=col_ind, end_row=(len2 + 1) + var2, end_column=col_ind)
        ws_summary.cell((len2+1) + var2+1,var4 + col_ind).value = f"=SUM({sum_col}{(len2+1)}:{sum_col}{(len2+1) + var2})"  # make summary of same city trunks in Minutes
        list2[3][1] = f"={sum_col}{(len2+1) + var2+1}"
        # simple border style set ranges
        cell_range_kh1 = [(len2 + 1), col_ind, (len2 + 1) + var2, var4 + col_ind]
        cell_range_kh2 = [(len2 + 1)+var2+1, var4 + col_ind, (len2 + 1) + var2+1, var4 + col_ind]

        # Total Domestic Time (Min) summation
        for itm,va1 in zip(list2,range(4)):
            for itm1,va2 in zip(itm,range(2)):
                ws_summary.cell(len3+va1, (col_ind + 2)+va2).value = itm1
        col_let = get_column_letter((col_ind + 2)+va2)
        ws_summary.cell(len3 + va1 + 1, (col_ind + 2) + va2).value = f"=SUM({col_let}{len3+va1-2}:{col_let}{len3+va1})"

        # final summary cells border styles ranges
        cell_range_su1 = [len3+1, (col_ind+2), len3 + va1, va2 + (col_ind+2)]
        cell_range_su2 = [len3+va1+1, va2+(col_ind+2), len3+va1+1, va2 + (col_ind+2)]
        cell_range_suH = [len3, (col_ind+2), len3, col_ind + 3]  # Starting column -3 means first col of summary header

        # set all border style to all
        list_bd_sty_H = [cell_range_dhH,cell_range_chH,cell_range_khH,cell_range_suH]   # set Header style of summary
        list_bd_sty = [cell_range_dh1,cell_range_ch1,cell_range_kh1,cell_range_su1]     # range of cell of data
        list_bd_sty2 = [cell_range_dh2,cell_range_ch2,cell_range_kh2,cell_range_su2]

        for bd_sty_r in list_bd_sty: self.set_border(ws_summary,cell_range=bd_sty_r,wraptext=False,font_size=14)
        for bd_sty_r in list_bd_sty2: self.set_border(ws_summary,cell_range=bd_sty_r,wraptext=False,font_size=14,font_weight=True)
        for bd_sty_r in list_bd_sty_H: self.set_border(ws_summary,cell_range=bd_sty_r,bd_sty=bd_sty,bg_color=header_color,font_weight=font_weight,font_size=14)

        # set Column width
        sd1 = [9]*(col_ind-1)+[18, 34, 32, 33, 9,9,9,9,9,9]
        for sd in range(1,ws_summary.max_column + 1):
            cd = ws_summary.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        # set height of the headers
        for itm in length: ws_summary.row_dimensions[itm].height = float(23.75)
        #  floating number will show up to two decimal point
        num_f = ws_summary.iter_rows(min_row=2, min_col=col_ind+3, max_row=ws_summary.max_row, max_col=ws_summary.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'
        return None

    def make_database_kpi_summary(self):
        global workbook_dataBase
        try:
            workbook_dataBase = load_workbook("workbook_dataBase.xlsx")
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]

            workbook_dataBase = Workbook()
            print(f"Exception {e} in loading excel. workbook_dataBase.xlsx File not found..")
            printCon(f"Exception {e} in loading excel. workbook_dataBase.xlsx File not found..", color='red')

        print(workbook_dataBase.sheetnames)
        printCon(workbook_dataBase.sheetnames)
        if "worksheet_kpi_summary_DB" not in workbook_dataBase.sheetnames:
            workbook_dataBase.create_sheet("worksheet_kpi_summary_DB")
            print(f"new sheet created.name:worksheet_kpi_summary_DB.............................info")
            printCon(f"new sheet created.name:worksheet_kpi_summary_DB.............................info")
        worksheet_kpi_summary_DB = workbook_dataBase["worksheet_kpi_summary_DB"]
        print(f"worksheet_ping = workbook_dataBase['worksheet_kpi_summary_DB'] created")


        month_number = self.db_var_summary_list[0].split("-")[1]
        month_date = self.db_var_summary_list[0].split("-")[2]
        year_no = self.db_var_summary_list[0].split("-")[0]
        month_number = int(month_number)-1
        month_date = int(month_date)
        year_no = int(year_no)
        if year_no < 2020: # no data acceptable before 2020...
            return False

        row = 10 + month_number * 32 + month_date + (year_no - 2020)*370
        col = 1
        for i, itm in zip(range(self.db_var_summary_list.__len__()),self.db_var_summary_list):
            worksheet_kpi_summary_DB.cell(row,col+i).value = itm

        sum1 = 0
        for itm2 in range(3):
            sum1 = int(self.db_var_summary_list[itm2+1]) + sum1
        worksheet_kpi_summary_DB.cell(row, col + i + 1).value = sum1

###########################################################################################################
        conn = sqlite3.connect("all_data.db")
        test = conn.cursor()
        test.execute("select summary_anable from  all_info where id = 1;")
        monthly_kpi_summary_enable = test.fetchone()[0]
        if monthly_kpi_summary_enable==1:
            # start making Monthly summary...
            self.wb_report.create_sheet(index=4,title="Monthly Summary")    # by default it creates with index=0
            Monthly_Summary = self.wb_report["Monthly Summary"]
            header = ["Date", "Dhaka Total Time (Min)", "CTG Total Time (Min)", "Khulna Total Time (Time)", "Total Domestic Time (Min)"]
            for item5 in range(5): Monthly_Summary.cell(2, 2 + item5).value = header[item5]

            item_ = 0
            for item in range(33):
                flag = True
                for item2 in range(5):
                    temp3 = worksheet_kpi_summary_DB.cell(row - item, 1 + item2).value
                    if not (temp3 == None or temp3 == "None" or temp3 == ""):
                        if item2 == 0:
                            Monthly_Summary.cell(3 + item_, 2 + item2).value = temp3
                        else:
                            temp3 = int(temp3)
                            Monthly_Summary.cell(3 + item_, 2 + item2).value = temp3/60
                        if flag and item2==4:
                            flag = False
                            item_ += 1
                    else:
                        pass

            # set header
            cell_range_cal = [2, 2, 2, 6]  # range of header of Monthly summary
            self.set_border(Monthly_Summary, cell_range=cell_range_cal, font_size=12, font_weight=True,
                            bd_sty="medium", bg_color=header_color,  wraptext=False)

            cell_range_cal = [3, 2, Monthly_Summary.max_row, 6]  # apply to all values
            self.set_border(Monthly_Summary, cell_range=cell_range_cal, wraptext=False)

            # set Column width
            sd1 = [5, 12, 25, 25, 27, 27,3]
            printCon(f"Set width of Excel file columns Monthly summary: {sd1}")
            for sd in range(1, Monthly_Summary.max_column + 1+1):
                cd = Monthly_Summary.column_dimensions[f"{get_column_letter(sd)}"]
                cd.width = sd1[sd - 1]
            rd = Monthly_Summary.row_dimensions[2]  # get dimension for row 1
            rd.height = 25  # value in points, there is no "auto"
            # Monthly_Summary.freeze_panes = "A3"  # make freeze before A2, means up to A1,B1,C1 . . .. .

            #  floating number will show up to two decimal point
            num_f = Monthly_Summary.iter_rows(min_row=3, min_col=3, max_row=Monthly_Summary.max_row, max_col=6)
            for num_1 in num_f:
                for num_2 in num_1:
                    num_2.number_format = '###0.00'

            from openpyxl.chart import (LineChart, Reference)
            from openpyxl.chart.axis import DateAxis
            from datetime import date

            data = Reference(Monthly_Summary, min_col=6, min_row=2, max_col=6, max_row=Monthly_Summary.max_row)

            # Chart with date axis
            chart1 = LineChart()
            chart1.width = 27
            chart1.height = 12
            chart1.title = "Monthly Summary"
            chart1.legend = None
            # chart1.legend.delete = True
            chart1.style = 1
            chart1.y_axis.title = "Total Domestic Time (Min)"
            # c2.y_axis.crossAx = 500
            # c2.x_axis = DateAxis(crossAx=100)
            chart1.x_axis.number_format = 'd-mmm'
            chart1.x_axis.majorTimeUnit = "days"
            chart1.x_axis.title = "Date"

            chart1.add_data(data, titles_from_data=True)
            dates = Reference(Monthly_Summary, min_col=2, min_row=3, max_row=Monthly_Summary.max_row,max_col=2)
            chart1.set_categories(dates)

            Monthly_Summary.add_chart(chart1, "H2")
###################################################################################3
        workbook_dataBase.save("workbook_dataBase.xlsx")
        printCon(f"Data base save successfully")
        print(f"Data base save successfully")

    def make_database_ios_summary(self):
        global workbook_dataBase
        try:
            workbook_dataBase = load_workbook("workbook_dataBase.xlsx")
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]

            workbook_dataBase = Workbook()
            print(f"Exception {e} in loading excel. workbook_dataBase.xlsx File not found..")
            printCon(f"Exception {e} in loading excel. workbook_dataBase.xlsx File not found..", color='red')

        print(workbook_dataBase.sheetnames)
        printCon(workbook_dataBase.sheetnames)
        if "worksheet_ios_summary_DB" not in workbook_dataBase.sheetnames:
            workbook_dataBase.create_sheet("worksheet_ios_summary_DB")
            print(f"new sheet created.name:worksheet_ios_summary_DB.............................info")
            printCon(f"new sheet created.name:worksheet_ios_summary_DB.............................info")
        worksheet_ios_summary_DB = workbook_dataBase["worksheet_ios_summary_DB"]
        print(f"worksheet_ping = workbook_dataBase['worksheet_ios_summary_DB'] created")

        print(self.db_var_summary_list.__len__(),self.db_var_summary_list)

        printCon(f"{self.db_var_summary_list.__len__()},{self.db_var_summary_list}")

        month_number = self.db_var_summary_list[0].split("-")[1]
        month_date = self.db_var_summary_list[0].split("-")[2]
        year_no = self.db_var_summary_list[0].split("-")[0]
        month_number = int(month_number) - 1
        month_date = int(month_date)
        year_no = int(year_no)
        if year_no < 2020:  # no data acceptable before 2020...
            return False

        row = 10 + month_number * 32 + month_date + (year_no - 2020) * 366
        col = 1

        for i, itm in zip(range(self.db_var_summary_list.__len__()), self.db_var_summary_list):
            worksheet_ios_summary_DB.cell(row, col + i).value = itm

        sum1 = 0
        for itm2 in range(1,self.db_var_summary_list.__len__()):
            sum1 = int(self.db_var_summary_list[itm2]) + sum1
        worksheet_ios_summary_DB.cell(row, col + i + 1).value = sum1//2

        workbook_dataBase.save("workbook_dataBase.xlsx")
        printCon(f"Data base save successfully")
        print(f"Data base save successfully")

    def summary_ios(self,all_dic):

        iso_summary = "IOS Summary"
        # remove sheet if exist already
        if iso_summary not in self.wb_report.sheetnames:
            self.wb_report.create_sheet(title=f"{iso_summary}",index=7)
        else:
            temp_ws = self.wb_report[iso_summary]
            self.wb_report.remove(temp_ws)
            self.wb_report.create_sheet(title=iso_summary,index=7)
        ws_summary = self.wb_report[iso_summary]

        # all_dic = list of dictionary of city trunks with value "Answer Time" ranges it will help to make sum.
        ios_trunk_list = ["1001-Roots_IOS_IN", "1005-NovoTel_IOS_IN", "1009-Btrac_IOS_IN", "1013-MirTelecom_IOS_IN", "1017-Global_Voice_IN", "1021-Unique_IOS","1025-Digicon_IOS"]
        ios_list = ["Roots", "NovoTel", "Btrac", "MirTelecom", "GlobalVoice", "Unique", "Digicon"]

        list1 =["IOS","Total Time (Sec)","Total Time (Min)"]
        list2 = ["IOS","Total International (Min)"]

        # variables that used
        col_ind = 7     # How many column will be empty before summary calculation
        len0 = 3        # position of Dhaka trunk summary
        len1 = len(ios_list) + len0 + 4   # CTG trunk summary starts
        length = [len0,len1]

        # Two hearer in excel
        # for itm,itm2 in zip(list1,range(len(list1))):   # print header of each summary
        for itm2,itm in enumerate(list1):   # print header of each summary
            ws_summary.cell(len0, col_ind+itm2).value = itm  # print header of each summary
        for itm2,itm in enumerate(list2):
            ws_summary.cell(len1, col_ind+itm2+1).value = itm

        #  set border style of header
        # Starting column -2 means first col of summary Header
        cell_range = []
        cell_range.append([len0,col_ind, len0,col_ind+2])
        # Minutes calculation
        cell_range.append([len1, (col_ind + 1), len1, col_ind + 2])
        for itm in range(2): self.set_border(ws_summary,cell_range[itm],bd_sty="medium",font_weight=True,bg_color=header_color,font_size=14)

        # ******************************************************************************************

        aa = all_dic[0].get(ios_trunk_list[0])
        print(aa)
        # exit()
        printCon(f"{all_dic}")
        for var1, var2 in zip(ios_list, range(len(ios_list))):
            list_dh = [var1, f"=SUM({var1}!{all_dic[var2].get(ios_trunk_list[var2])})", f"=SUM({var1}!{all_dic[var2].get(ios_trunk_list[var2])})/60"]

            for var3, var4 in zip(list_dh, range(3)):
                ws_summary.cell((len0+1) + var2, var4 + col_ind).value = var3

            cell_range = [(len0+1) + var2, col_ind,(len0+1) + var2, var4 + col_ind]
            self.set_border(ws_summary,cell_range,font_size=14,wraptext=False)
        cell_range = [(len0 + 1), col_ind, (len0 + 1) + var2, col_ind]  # Cell range of IOS name to make font weight bold
        self.set_border(ws_summary,cell_range,wraptext=False,font_size=14,font_weight=True)

        # Total isd Time (Min) summation ************************
        for itm,va1 in zip(ios_list,range(len(ios_list))):
            sum_list = [f"{itm}",   f"=SUM({itm}!{all_dic[va1].get(ios_trunk_list[va1])})/60"]

            for itm1,va2 in zip(sum_list,range(2)):
                ws_summary.cell(len1+va1+1, (col_ind + 1)+va2).value = itm1

            # set all cell property/style
            cell_range = [len1+va1+1, (col_ind + 1), len1+va1+1, (col_ind + 1) + va2]
            self.set_border(ws_summary,cell_range,wraptext=False,font_size=14)

        col_let = get_column_letter((col_ind + 1)+va2)      # for sum calculation
        ws_summary.cell(len1 + va1 + 2, (col_ind + 1) + va2).value = f"=SUM({col_let}{len1+1}:{col_let}{len1+va1+1})"
        # print(f"=SUM({col_let}{len1+1}:{col_let}{len1+va1+1})")
        printCon(f"=SUM({col_let}{len1+1}:{col_let}{len1+va1+1})")

        #   set final sum only one cell Style
        cell_range = [len1 + va1 + 2, (col_ind + 1) + va2, len1 + va1 + 2, (col_ind + 1) + va2]      # for summary cell, only one cell (va1 + 2) = 9
        self.set_border(ws_summary,cell_range,wraptext=False,font_weight=True,font_size=14)

        #   set style left side of Total sum
        cell_range = [(len1 + 1), col_ind+1, (len1 + 1) + var2, col_ind+1]  # Cell range of IOS name to make font weight bold
        self.set_border(ws_summary, cell_range, wraptext=False, font_size=14, font_weight=True)

        # set Column width
        sd1 = [9]*(col_ind-1)+[28, 32, 32, 9, 9, 9, 9, 9, 9, 9]
        for sd in range(1,ws_summary.max_column + 1):
            ws_summary.column_dimensions[f"{get_column_letter(sd)}"].width = sd1[sd - 1]

        # set height of the headers
        for itm in length: ws_summary.row_dimensions[itm].height = float(27)

        #  floating number will show up to two decimal point
        num_f = ws_summary.iter_rows(min_row=len0, min_col=col_ind+2, max_row=ws_summary.max_row, max_col=col_ind+2)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        return None

    def set_border(self, ws, cell_range,bd_sty="thin",font_weight=False,bg_color = "ffffff",wraptext = True,font_size = 11):  # Function to set cell properties
        #bd_sty = "thin"
        bd_color = "000000"
        border = Border(
            left=Side(border_style=bd_sty, color=bd_color),
            right=Side(border_style=bd_sty, color=bd_color),
            top=Side(border_style=bd_sty, color=bd_color),
            bottom=Side(border_style=bd_sty, color=bd_color)
        )  # this indentation helps to comment unnecessary command line.
        font = Font(name="Times New Roman",size=font_size,bold=font_weight,color="000000")
        # set font bold,italic are boolen, font color etc
        alignment = Alignment(horizontal="center",vertical="center",wrapText=wraptext)
        patternfill = PatternFill(start_color=bg_color,end_color=bg_color, fill_type="solid")  # set cell background color
        # print("Cell range from cell property method ",cell_range)
        # printCon(f"Set Style; Cell range from cell property method: {cell_range}; thread name:  {threading.current_thread().getName()}")
        rows = ws.iter_rows(min_row=cell_range[0],min_col=cell_range[1],max_row=cell_range[2],max_col=cell_range[3])  # iter_rows function make a tuple of tuples of cell objects.
        for row in rows:
            for cell in row:
                cell.border = border        # set Border properties
                cell.font = font            # set font styles
                cell.alignment = alignment  # set alignment
                cell.fill = patternfill     # set background color.
        return ws

    def make_IDD_report(self):
        global kpi_top
        global text_var_out_folder
        try:
            fd1 = open(self.filename, "r");
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("Please in put file")
            printCon(f"Please in put IDD file", color='red')
            # text_var_out_folder.set("Please in put first.")
            messagebox.showwarning("warning", "Invalid Input\nPlease input first.", parent=kpi_top)
            return None

        if len(fd1.readline().split('","')) != 41:
            # text_var_out_folder.set("Wrong in put file.")
            messagebox.showwarning("warning", "Wrong input file.", parent=kpi_top)
            return None

        fd1.seek(0, 0);
        length_fd1 = len(fd1.readlines());
        fd1.seek(0, 0);
        wb_idd = Workbook();
        sh_list = wb_idd.sheetnames;
        sh_list[0] = "idd raw";
        sh_list = wb_idd.sheetnames;
        ws_idd = wb_idd[sh_list[0]];
        fd1.seek(0, 0);
        A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25
        del_col = [A] * 6 + [B] * 2 + [C]*3 + [D, H] + [O] * 3 + [R]*7  # to be delete column list
        printCon(f"Column to be deleted: {del_col}")
        fd1.seek(0, 0);  # this command take file pointer at 0,0 position
        for itm in range(length_fd1):  # this loop work at the end of the line  of csv file
            row_list = fd1.readline().split('","');
            for item in del_col:  # delete unnecessary column from .csv file
                del row_list[item];
            row_list[-1] = row_list[-1][:-3]  # remove last list unnecessary ", sign which was created in csv file
            row_list.insert(0, " ")  # insert first space for indent
            for item1 in range(len(row_list)):  # Convert string to integer
                if row_list[item1].isdecimal():
                    row_list[item1] = int(row_list[item1])
            if itm == 0:
                index_connect_number = row_list.index("Connect Number")
                index_attempt_number = row_list.index("Attempt Number")
                index_Answer_Number = row_list.index("Answer Number")
                index_Answer_Time = row_list.index("Answer Time")
                row_list.append("ASR")
                row_list.append("ACD")
                row_list.append("CCR")
            else:
                max_ro = itm;
                row_list.append(
                    f"={chr(65 + index_Answer_Number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")
                row_list.append(
                    f"={chr(65 + index_Answer_Time)}{max_ro + 1}/{chr(65 + index_Answer_Number)}{max_ro + 1}/60")
                row_list.append(
                    f"={chr(65 + index_connect_number)}{max_ro + 1}/{chr(65 + index_attempt_number)}{max_ro + 1}*100")

            ws_idd.append(row_list);
        fd1.close();
        min_row = ws_idd.min_row
        min_col = ws_idd.min_column
        max_row = ws_idd.max_row
        max_col = ws_idd.max_column
        # col_letter = get_column_letter(max_col)

        ''''
        Setting style to the border and fonts, 
        '''
        cell_range = [min_row, min_col + 1, max_row, max_col]  # set style all sheet.
        self.set_border(ws_idd, cell_range)

        top_cell_range = [1, 2, 1, ws_idd.max_column]  # set style top row/ Header row
        self.set_border(ws_idd, top_cell_range, bd_sty="medium", font_weight=True,
                        bg_color=header_color)  # Border width = medium valid !!

        top_cell_range = [2, ws_idd.max_column - 2, ws_idd.max_row, ws_idd.max_column]  # set style ASR,ACD,CCR
        self.set_border(ws_idd, top_cell_range, bd_sty="thin", font_weight=True, bg_color=side_color)

        rd = ws_idd.row_dimensions[1]  # get dimension for row 3
        rd.height = 48  # value in points, there is no "auto"
        # sd1 = [16,14,9,12,15,15,12,10,10,11,11,10M,12,11,11,13Q,11,11,11]
        sd1 = [4, 16, 14, 13, 15, 15, 12, 10, 11, 11, 12, 12, 13, 12, 12, 13, 13, 13, 12, 9, 9, 9,9]
        for sd in range(1, ws_idd.max_column + 1):
            cd = ws_idd.column_dimensions[f"{get_column_letter(sd)}"]
            cd.width = sd1[sd - 1]
        ws_idd.freeze_panes = "A2"  # make freeze before B2
        # ws_idd.insert_cols(1)  # this will insert correctly but little problem with my fix excel formulas

        #  floating number will show up to two decimal point
        num_f = ws_idd.iter_rows(min_row=2, min_col=ws_idd.max_column - 2, max_row=ws_idd.max_row,
                                 max_col=ws_idd.max_column)
        for num_1 in num_f:
            for num_2 in num_1:
                num_2.number_format = '#,##0.00'

        # save file with name
        # set output folder
        if os.path.isdir(out_folder):
            pass
        else:
            Options.change_out_folder()

        # Change output folder
        global out_folder_dir
        try:
            os.chdir(out_folder_dir)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]

        x = time.localtime(os.path.getctime(self.filename))
        date_file = ws_idd.cell(2, 3).value.split("-")
        month_file = month_name[int(date_file[1])]

        if x[3] > 12:        # for morning report it is false.
            time_upto = datetime.strptime(f"{x[3]}", "%H").strftime("%I %p")
            name = f"IDD Report {date_file[2]} {month_file} {date_file[0]} (Upto {time_upto}).xlsx"
            time_upto = f" up to {time_upto}"
        else:
            name = f"IDD Report {date_file[2]} {month_file} {date_file[0]}.xlsx"
            time_upto = ''

        wb_idd.save(name)

        # resave the Excel file with MS Excel to make compatible with Excel in mail. or format conversion
        path1 = os.getcwd() + os.sep
        path2 = os.path.join(path1, name)
        try:
            xl = Dispatch("Excel.Application")
            wb2 = xl.Workbooks.Open(Filename=path2)
            xl.Visible = False  # speed up process also
            wb2.Save()  # Save and over lap the original file
            wb2.Close(True)
            xl.Quit()
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"Exception type: {e}\nMS Excel may be not installed in this system...")
            printCon(f"Exception type: {e}\nMS Excel may be not installed in this system...", color='red')
        # set current directory as previous
        os.chdir(current_working_dir)

        # Print completion message
        kip_conform.set(f"IDD complete morning / Evening")
        # except Exception: print("File not found \n Or Wrong file selected \nOut file has opened")

        # for Mailing purpose
        mmcl_domestic_ISD_report_class.nam_temp_idd = name
        mmcl_domestic_ISD_report_class.nam_temp_idd_t = [date_file[0], month_file, date_file[2], time_upto]
        if auto_mail_enable.get() == 1:
            mail_instance.idd()

        return None

    print("")

class Open_gui:

    @classmethod
    def menu_bars(cls):
        # Menu make
        menubar = Menu(top)
        file = Menu(menubar,tearoff=False)
        file.add_command(label="Set Header Color", command=Options.headerColor)
        file.add_command(label="Set Side color", command=Options.sideColor)
        file.add_command(label="Current Settings", command=Options.current_d)
        file.add_command(label="Change Output folder", command=Options.change_out_folder)
        file.add_separator()
        file.add_command(label="Exit", command=extraInfo.callback)
        menubar.add_cascade(label="File", menu=file)

        tools = Menu(menubar,tearoff=False)
        tools.add_command(label="Digital Clock", command=Tools.d_clock)
        tools.add_command(label="KPI, IOS_ISD, IDD Report", command=Open_gui.kpi_gui)
        tools.add_command(label="SPC Converter", command=Tools.spc_converter)
        ping = Menu(tools,tearoff=False)
        ping.add_command(label="Open All Ping", command=Tools.cmd)
        ping.add_command(label="All Link Status", command=Tools.cmdSetting)
        ping.add_command(label="Ping Setting", command=partial(ping_setting.ping_set.create_Toplevel1,top))
        tools.add_cascade(menu=ping, label="Ping")

        global auto_mail_enable
        auto_mail_enable = IntVar();auto_mail_enable.set(0)
        def update_enable_mail_Var():
            if auto_mail_enable.get() == 1:
                ato = "Enable"
            else: ato = "Disable"
            text_var_auto_mail_e.set(ato)
        tools.add_checkbutton(label="Auto Mail Enable",variable = auto_mail_enable, offvalue=0,
                              command=update_enable_mail_Var,onvalue=1)
        menubar.add_cascade(menu=tools, label="tools")

        global mail_instance
        mail_instance = Mail("Md. Habibur Rahman")
        mail_menu = Menu(menubar,tearoff=False)
        mail_menu.add_command(label="create new Mail", command=mail_instance.creat_mail)
        mail_menu.add_command(label="Peak Hour Traffic Statistics", command=mail_instance.Peak_Hour_Traffic_Statistics)
        mail_menu.add_command(label="IDD Every Two Hours", command=mail_instance.idd_every_two_hours)
        mail_menu.add_command(label="KPI report", command=mail_instance.kpi)
        mail_menu.add_command(label="IOS ISD report", command=mail_instance.ios_idd)
        mail_menu.add_command(label="IDD report", command=mail_instance.idd)
        menubar.add_cascade(menu=mail_menu, label="Mail")

        noc_name = Menu(menubar,tearoff=False)
        noc_name.add_radiobutton(label="Habib", command=partial(mail_instance.__init__, "Md. Habibur Rahman"))
        noc_name.add_radiobutton(label="Turzo", command=partial(mail_instance.__init__, "Akil Monsur"))
        noc_name.add_radiobutton(label="Turaz", command=partial(mail_instance.__init__, "Tanzil Monsur"))
        noc_name.add_radiobutton(label="Amit", command=partial(mail_instance.__init__, "Amit Roy"))
        noc_name.add_radiobutton(label="Tusher", command=partial(mail_instance.__init__, "Md. Tasnim Rahman Tusher"))
        noc_name.add_radiobutton(label="Obeyddullah", command=partial(mail_instance.__init__, "A.E.M Obeyddullah Siddique"))
        menubar.add_cascade(menu=noc_name, label="Employee")

        Help = Menu(menubar,tearoff=False)
        Help.add_command(label="Document",command=Tools.Help_doc_login)
        Help.add_command(label="Help",command=extraInfo.show_image_doc)
        Help.add_command(label="Formulas",command=partial(formula_gui.formula_.create_Toplevel1, top))
        Help.add_command(label="show console", command=partial(console_out.console_out_gui_file.create_Toplevel1, top))
        Help.add_command(label="About Me",command=extraInfo.about)
        menubar.add_cascade(menu=Help, label="Document")

        top.config(menu=menubar)

        # # Excel Header color set
        # exists = os.path.isfile(r'data/color/header_color.txt')
        # global header_color,side_color
        # if exists:
        #     header_color = open(r'data/color/header_color.txt', "r").read(6)
        # else:
        #     header_color = "ff0000"
        #
        #     # Excel side  color set
        # exists = os.path.isfile(r'data/color/side_color.txt')
        # if exists:
        #     side_color = open(r'data/color/side_color.txt', "r").read(6)
        # else:
        #     side_color = "00ff00"
        # return None

    global kpi_top_control_bool
    kpi_top_control_bool = False

    @classmethod
    def kpi_gui(cls):
        global top, kpi_top_control_bool
        if kpi_top_control_bool:
            printCon(f"You attempt to Open 'KPI, IOS_ISD, IDD Report' while it is already running..",color='purple')
            return None
        kpi_top_control_bool = True
        def kpi_top_control_func():
            global kpi_top_control_bool
            kpi_top_control_bool = False
            # top.wm_state('normal')
            kpi_top.destroy()
        '''
        Starting point of periodic monitoring
        '''
        # top.wm_state('iconic') # used to minimize the main window...
        global kpi_top
        kpi_top = Toplevel()
        kpi_top.protocol("WM_DELETE_WINDOW", kpi_top_control_func) #catch close command when click crose
        kpi_top.resizable(width=False,height=False)
        kpi_top.title("KPI, IOS_ISD, IDD Report")

        menubar = Menu(top)
        file = Menu(menubar,tearoff=False)
        file.add_command(label="Set Header Color", command=Options.headerColor)
        file.add_command(label="Set Side color", command=Options.sideColor)
        file.add_command(label="Current Settings", command=Options.current_d)
        file.add_command(label="Change Output folder", command=Options.change_out_folder)
        file.add_separator()
        # def des():
        #     kpi_top.destroy()
        #     top.wm_state('normal')
        file.add_command(label="Exit", command=kpi_top_control_func)
        menubar.add_cascade(label="File", menu=file)

        global mail_instance
        mail_menu = Menu(menubar,tearoff=False)
        mail_menu.add_command(label="create new Mail", command=mail_instance.creat_mail)
        mail_menu.add_command(label="IDD Every Two Hours", command=mail_instance.idd_every_two_hours)
        mail_menu.add_command(label="KPI report", command=mail_instance.kpi)
        mail_menu.add_command(label="IOS ISD report", command=mail_instance.ios_idd)
        mail_menu.add_command(label="IDD report", command=mail_instance.idd)
        menubar.add_cascade(menu=mail_menu, label="Mail")

        noc_name = Menu(menubar,tearoff=False)
        noc_name.add_radiobutton(label="Habib", command=partial(mail_instance.__init__, "Md. Habibur Rahman"))
        noc_name.add_radiobutton(label="Turzo", command=partial(mail_instance.__init__, "Akil Monsur"))
        noc_name.add_radiobutton(label="Turaz", command=partial(mail_instance.__init__, "Tanzil Monsur"))
        noc_name.add_radiobutton(label="Amit", command=partial(mail_instance.__init__, "Amit Roy"))
        noc_name.add_radiobutton(label="Tusher", command=partial(mail_instance.__init__, "Md. Tasnim Rahman Tusher"))
        noc_name.add_radiobutton(label="Obeyddullah", command=partial(mail_instance.__init__, "A.E.M Obeyddullah Siddique"))
        menubar.add_cascade(menu=noc_name, label="Employee")

        kpi_top.config(menu=menubar)

        bg_kpi1 = "#a2d32c"
        bg_kpi2 = "#99aa00"
        Label(kpi_top).grid(row=1, column=0)  # gap between two part

        global raw_kpi_in_instance,raw_kpi_out_instance,raw_idd
        raw_kpi_in_instance = mmcl_domestic_ISD_report_class();
        raw_kpi_out_instance = mmcl_domestic_ISD_report_class();
        raw_idd = mmcl_domestic_ISD_report_class();

        frm_kpi = Frame(kpi_top, bd=5, relief="solid", pady=5, padx=5, bg="#ffffff")
        raw_frame = Frame(frm_kpi, bd=5, relief="solid", width=500, bg="#ff009f");
        make_frame = Frame(frm_kpi, bd=5, relief="solid", width=500, bg="#ff009f");

        global raw_in_kpi_file_path,raw_out_kpi_file_path,raw_idd_file_path
        raw_in_kpi_file_path = StringVar();
        raw_in_kpi_file_path.set("Open incoming KPI .csv")
        raw_out_kpi_file_path = StringVar();
        raw_out_kpi_file_path.set("Open outgoing KPI .csv")
        raw_idd_file_path = StringVar();
        raw_idd_file_path.set("Open IDD .csv")

        btn_1 = Button(raw_frame, command=partial(raw_kpi_in_instance.upload_kpi,kpi_top), textvariable=raw_in_kpi_file_path,
                       width=width_1, font="Times 15", bg=bg_kpi1, anchor="w",relief=relief,bd=bd);
        btn_2 = Button(raw_frame, command=partial(raw_kpi_out_instance.upload_kpi_out,kpi_top), textvariable=raw_out_kpi_file_path,
                       width=width_1, font="Times 15", bg=bg_kpi1,relief=relief,bd=bd, anchor="w");
        btn_3 = Button(raw_frame, command=partial(raw_idd.upload_idd,kpi_top), textvariable=raw_idd_file_path, width=width_1,
                       font="Times 15", bg=bg_kpi1,relief=relief,bd=bd, anchor="w");

        Label(frm_kpi, width=width_1, font="Times 6").grid(row=3, column=0)  # gep between row file and make report

        width_2 = round(width_1 / 3) - 1
        btn_kpi = Button(make_frame, text="Make KPI Report", command=AbstractClass.in_comming_kpi, width=width_2,
                         relief=relief, bd=bd, font=font, bg=bg_kpi2)
        btn_kpi.grid(row=0, column=0);

        btn_ios = Button(make_frame, text="Make IOS IDD Report", command=AbstractClass.out_going_kpi,
                         relief=relief, bd=bd, width=width_2, font=font, bg=bg_kpi2)
        btn_ios.grid(row=0, column=1);

        btn_idd = Button(make_frame, text="Make IDD Report", command=AbstractClass.idd_day_calculation_time_period, width=width_2, font=font,
                         relief=relief, bd=bd, bg=bg_kpi2)
        btn_idd.grid(row=0, column=2);

        btn_1.grid(row=1, column=1);
        btn_2.grid(row=2, column=1);
        btn_3.grid(row=3, column=1);

        raw_frame.grid(row=2, column=0);  # row=2 in main frame "frm_kpi"
        make_frame.grid(row=4, column=0);  # row=4 in main frame "frm_kpi"
        frm_kpi.grid(row=3, column=0)  # this is the main frame for KPI report, inside top

        # show conformation messege
        global kip_conform
        kip_conform = StringVar()
        kip_conform.set(" ")
        Label(frm_kpi, textvariable=kip_conform, font=font, width=width_1,bg="white").grid(row=5, column=0)

        # Exit button
        Button(frm_kpi,text='Exit',font=font,command=kpi_top_control_func,width=width_1,relief="raised",
               bg="#fab005",bd=10).grid()
        return None

    @classmethod
    def periodic_gui(cls):
        pass

class Tools:

    def __init__(self):
        pass

    @classmethod
    def d_clock(cls):
        root = Toplevel()
        root.title("Digital Clock")
        clock = Label(root, font="Times 100",bg="white")
        root.resizable(width=False,height=False)
        clock.pack()
        def clock_refresher():
            time_string = time.strftime("%H:%M:%S")
            clock.config(text=time_string)
            clock.after(50, clock_refresher)
        Thread(target=clock_refresher).start()
        return None

    @classmethod
    def spc_converter(cls):
        width1 = 15
        width2 = 25

        top_spc = Toplevel()

        width_of_window = 465;
        height_of_window = 290;
        screen_width = top.winfo_screenwidth();
        screen_height = top.winfo_screenheight();
        x_coordinate = screen_width / 2 - width_of_window / 2;
        y_coordinate = screen_height / 2 - height_of_window / 2;
        top_spc.geometry("%dx%d+%d+%d" % (width_of_window, height_of_window, x_coordinate, y_coordinate))

        top_spc.resizable(0,0)
        top_spc.title("SPC Conversion")
        Label(top_spc,text="SS7 Point Code Converter",font=font,bg="#aabbaa",width=width1+width2+1
              ,height=2).grid()

        frame_top1 = Frame(top_spc,bd=bd,relief="solid")
        def calculate():
            a = StVar.get().split("-")
            try:
                a=[int(itm) for itm in a]
                if len(a) != 3:
                    raise IOError
                elif a[1]>255:
                    messagebox.showwarning("warning", "Invalid Input\nMax Cluster Number 255", parent=top_spc)
                elif a[2]>7:
                    messagebox.showwarning("warning", "Invalid Input\nMax Member Number 7", parent=top_spc)
                elif a[0]>7:
                    messagebox.showwarning("warning", "Invalid Input\nMax Network Number 7", parent=top_spc)
            except Exception as e:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                traceback.print_exception(exc_type, exc_value, exc_traceback)
                exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
                [printCon(i, color='red') for i in exc]
                messagebox.showwarning("warning", "Invalid Input\nEnter 3-8-3 Format.", parent=top_spc)
                return None

            c=[2048,8,1]
            dec_spc=0
            for itm in range(3):
                b=0
                try:
                    b = a[itm]*c[itm]
                except Exception as e:
                    exc_type, exc_value, exc_traceback = sys.exc_info()
                    traceback.print_exception(exc_type, exc_value, exc_traceback)
                    exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
                    [printCon(i, color='red') for i in exc]
                    print("Error type: ",e)
                    messagebox.showwarning("warning","Invalid Input\nEnter 3-8-3 Format",parent=top_spc)
                    return None
                dec_spc += b
            print(dec_spc)
            printCon(dec_spc)
            StVar_result.set(dec_spc)
        Label(frame_top1,text="SPC 3-8-3",font=font,width=width1).grid(row=1,column=0)
        StVar = StringVar()
        Entry(frame_top1,textvariable=StVar,font=font,width=width2).grid(row=1,column=1)
        Button(frame_top1,font=font,text="Calculate Decimal",command=calculate,width=width1,
               relief=relief,bd=bd,bg='#9f8c60').grid(row=2,column=0)
        StVar_result = StringVar()
        Label(frame_top1,font=font,textvariable=StVar_result,width=width2).grid(row=2,column=1)
        frame_top1.grid(row=1,column=0)

        Label(top_spc,width=width2+width1+1,bg="#aabb99",font=font).grid(row=2,column=0)

        frame_top2 = Frame(top_spc,bd=bd,relief="solid")
        def calculate_2():
            a = StVar_2.get()
            try:
                if int(a)>16383:messagebox.showwarning("warning", "Invalid Input\nMax Decimal Number 16383",parent=top_spc)
                b1 = '000000000000'+bin(int(a))[2:]
            except Exception as e:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                traceback.print_exception(exc_type, exc_value, exc_traceback)
                exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
                [printCon(i, color='red') for i in exc]
                print(f"Error type : ,{e}")
                printCon(f"Error type : ,{e}", color='red')
                messagebox.showwarning("warning", "Invalid Input\nEnter Decimal Number",parent=top_spc)
                return None
            b2 = [b1[-14:-11], b1[-11:-3], b1[-3:]]
            b3 = [int(b2[itm],2) for itm in range(3)]
            StVar_result_2.set(f'{b3[0]}-{b3[1]}-{b3[2]}')

        Label(frame_top2,text="SPC Decimal",font=font,width=width1).grid(row=1,column=0)
        StVar_2 = StringVar()
        Entry(frame_top2,textvariable=StVar_2,font=font,width=width2).grid(row=1,column=1)
        Button(frame_top2,font=font,text="Calculate 3-8-3",command=calculate_2,width=width1,
               relief=relief,bd=bd,bg='#9f8c60').grid(row=2,column=0)
        StVar_result_2 = StringVar()
        Label(frame_top2,font=font,textvariable=StVar_result_2,width=width2).grid(row=2,column=1)
        frame_top2.grid(row=3,column=0)

        Button(top_spc,font=font,command=top_spc.destroy,text="Exit",width=width2+width1+1,bg="#cccccc"
               ,height=2).grid()
        top_spc.mainloop()

    @classmethod
    def Help_doc_login(cls):

        def spc():
            var_spc_converter_Doc1 = """
            
            How do I perform point code conversion for SS7?
            
            The SS7 firmware requires point codes to be implemented in their Decimal form. 
            If you are given a point code in the Network-Cluster-Member form you will have to convert it to the Decimal form before downloading the firmware.
            
            Decimal point codes can range from 0 to 16383.
            
            Network-Cluster-Member point codes can range from: 0 to 7 (Network) - 0 to 255 (Cluster) - 0 to 7 (Member)
            
            Converting from Network-Cluster-Member to Decimal Form
            The function below can be used to perform a Network-Cluster-Member to Decimal conversion:
            
            Decimal Point Code = (2048 x Network) + (8 x Cluster) + Member
            
            Let's look at a worked example, take the Point Code 3-115-6.
            
            Thus,    Network = 3
            Cluster = 115
            Member = 6
            
            Decimal Point Code = (2048 x 3) + (8 x 115) + 6
            
            Decimal Point Code = 07070
            
            Converting from Decimal to Network-Cluster-Member form
            
            Converting from the Decimal form to the Network-Cluster-Member form can be achieved in three simple steps;
            
            Step 1: Convert the Decimal point code into Binary
            
            07070 (decimal) = 1B9E (hex) = 0001 1011 1001 1110 (binary)
            
            Step 2: Split the binary number into three parts, bits 1 to 3, bits 4 to 11 and bits 12 to 14.
            
            Bits 1 to 3 relate to the Member, bits 4 to 11 relate to the Cluster and bits 12 to 14 relate to the Network.
            (Network) 	(Cluster) 	(Member)
            bits 12 to 14 	bits 4 to 11 	bits 1 to 3
            011 	01110011 	110
            
            Step3: Finally convert the binary parts to decimal values.
            Network 	011 (binary) 	3 (decimal)
            Cluster 	01110011 (binary) 	115 (decimal)
            Member 	110 (binary) 	6 (decimal)
            
            Hence, Network-Cluster-Member form = 3 -115-6
            
             
            
            Categories:  Firmware Non-cloud SS7

            """


            top = Toplevel()
            top.geometry('%dx%d'%(int(top.winfo_screenwidth()*.9),int(top.winfo_screenheight()*0.7)))
            top.title("Point code conversion for SS7")
            text = Text(top)
            vs = Scrollbar(top, orient="vertical")
            hs = Scrollbar(top, orient="horizontal")
            sizegrip = ttk.Sizegrip(top)

            # hook up the scrollbars to the text widget
            text.configure(yscrollcommand=vs.set, xscrollcommand=hs.set, wrap="none")
            vs.configure(command=text.yview)
            hs.configure(command=text.xview)

            # grid everything on-screen
            text.grid(row=0, column=0, sticky="news")
            vs.grid(row=0, column=1, sticky="ns")
            hs.grid(row=1, column=0, sticky="news")
            sizegrip.grid(row=1, column=1, sticky="news")
            top.grid_rowconfigure(0, weight=1)
            top.grid_columnconfigure(0, weight=10)

            text.insert("end",var_spc_converter_Doc1)

            top.mainloop()
            return None

        def show_soruce_code():
            top = Toplevel()
            top.geometry('%dx%d' % (int(top.winfo_screenwidth() * .9), int(top.winfo_screenheight() * 0.7)))
            top.title("Source Code")
            text = Text(top)
            vs = Scrollbar(top, orient="vertical")
            hs = Scrollbar(top, orient="horizontal")
            sizegrip = ttk.Sizegrip(top)

            # hook up the scrollbars to the text widget
            text.configure(yscrollcommand=vs.set, xscrollcommand=hs.set, wrap="none")
            vs.configure(command=text.yview)
            hs.configure(command=text.xview)

            # grid everything on-screen
            text.grid(row=0, column=0, sticky="news")
            vs.grid(row=0, column=1, sticky="ns")
            hs.grid(row=1, column=0, sticky="news")
            sizegrip.grid(row=1, column=1, sticky="news")
            top.grid_rowconfigure(0, weight=1)
            top.grid_columnconfigure(0, weight=10)

            # first print without line Number
            text.insert("end", open(__file__).read())   # source code print in Text

            # second print with line Number Notation
            fk = open(__file__)
            ab = len(fk.readlines())
            fk.seek(0,0)
            for itm in range(ab):
                text.insert("end", f"{itm+1}{fk.readline()}")   # source code print in Text

            top.mainloop()
            return None

        width_label = 20
        width_entry = 26
        login_window = Toplevel()
        login_window.title("log in")
        login_window.resizable(0,0)
        width_of_window = 832
        height_of_window = 377
        screen_width = login_window.winfo_screenwidth()
        screen_height = login_window.winfo_screenheight()
        x_coordinate = screen_width / 2 - width_of_window / 2
        y_coordinate = screen_height / 2 - height_of_window / 2
        login_window.geometry("%dx%d+%d+%d" % (width_of_window, height_of_window, x_coordinate, y_coordinate))

        str_var = StringVar()
        frm_log_main = Frame(login_window, bg="gray")   # main window
        frm_log_main2 = Frame(login_window) # Doc menu page
        frm_log_main3 = Frame(login_window) # reset password window

        def login():
            nonlocal str_var
            temp1 = str_var.get()
            print(temp1)
            printCon(temp1)
            temp2 = hashlib.md5(temp1.encode())
            password1 = temp2.hexdigest()

            # make required folders
            cwd = os.getcwd()  # reserve current directory
            if not os.path.exists('data'):
                os.mkdir('data')
            os.chdir("data")
            if not os.path.exists('password'):
                os.mkdir('password')
            os.chdir(cwd)  # make current directory as previous

            conn = sqlite3.connect("all_data.db")
            cursor = conn.cursor()
            cursor.execute("select password from all_info where id = 1")
            passH = cursor.fetchone()[0]
            print(passH," ppppppppppppppppppppppppppppppp")
            # if not os.path.exists(r'data/password/password.txt'):
            if passH == 'password':
                h = hashlib.md5("123456".encode())
                h=h.hexdigest()
                conn.execute(f"update all_info set password = '{h}' where id=1;")
                conn.commit()
                # file_1 = open(r'data/password/password.txt', "w+")
                # file_1.write(str(h))
                print(f"password file created {h}")
                printCon(f"password file created, password = {h}")


            # file_2 = open(r'data/password/password.txt', "r")
            # password2 = file_2.read()
            cursor.execute("select password from all_info where id = 1")
            password2 = cursor.fetchone()[0]
            conn.close()
            printCon(f"{password1},{type(password1)}")
            print(password2,type(password2))
            printCon(f"password2 {password2},password2 {type(password2)}")
            if password1 == password2:
                # messagebox.showinfo("Log In", "Login Successful", parent=login_window)
                already_loged_in = 1
                doc_page()
                return True
            else:
                messagebox.showwarning("warning", "Wrong Password", parent=login_window)
                return False

            return False

        def reset_password():
            def reset_p():
                # fd1 = open(r'data/password/password.txt', "r")
                # password1 = fd1.read()

                conn = sqlite3.connect("all_data.db")
                cursor = conn.cursor()
                cursor.execute("select password from all_info where id = 1")
                password1 = cursor.fetchone()[0]
                # conn.close()
                password2=ent_old_pass.get()
                password2 = hashlib.md5(password2.encode())
                password2 = password2.hexdigest()

                if password1 == password2:
                    h = hashlib.md5(ent_new_pass.get().encode())
                    h = h.hexdigest()
                    conn.execute(f"update all_info set password = '{h}' where id = 1")
                    conn.commit()
                    # file_1 = open(r'data/password/password.txt', "w")
                    # file_1.write(str(h))
                    messagebox.showinfo("Info", "Password Reset Successfully", parent=login_window)
                else:
                    messagebox.showinfo("Info", "Something Wrong", parent=login_window)
                return None

            frm_log_main.destroy()  # to destroy the login page
            frm_log_main3.grid(row=0,column=0)

            frm_reset1 = Frame(frm_log_main3)
            frm_reset1.grid(row=0,column=0)

            lab_old_pass = Label(frm_reset1,text="Enter Old Password",width=width_label,font=font)
            lab_new_pass = Label(frm_reset1,text="Enter New Password",width=width_label,font=font)
            str_var_old_p = StringVar()
            str_var_new_p = StringVar()
            ent_old_pass = Entry(frm_reset1,width=width_entry+3,font=font,textvariable=str_var_old_p,show="*")
            ent_new_pass = Entry(frm_reset1,width=width_entry+3,font=font,textvariable=str_var_new_p,show="*")

            frm_reset2 = Frame(frm_log_main3)
            frm_reset2.grid(row=1,column=0)
            btn_reset = Button(frm_reset2,text="Reset password",font=font,width=(width_label+width_entry)//2,command=reset_p)
            btn_back = Button(frm_reset2,text="Back to Login",font=font,width=(width_label+width_entry)//2,command=main_frm)
            btn_exit = Button(frm_log_main3,text="Exit",font=font,width=(width_label+width_entry+1),bg="yellow",
                              command=login_window.destroy)

            ent_old_pass.grid(row=0,column=1)
            ent_new_pass.grid(row=1,column=1)
            lab_old_pass.grid(row=0,column=0)
            lab_new_pass.grid(row=1,column=0)
            btn_reset.grid(row=0,column=0)
            btn_back.grid(row=0,column=1)
            btn_exit.grid(row=2,column=0)

            return None

        def doc_page():
            nonlocal frm_log_main2,frm_log_main
            frm_log_main.destroy()  # to destroy the login page

            frm_log_main2.grid(row=0,column=0)
            Label(frm_log_main2,text="new window !!!").grid(row=0,column=0)

            btn_spc_doc = Button(frm_log_main2,text="SPC Converter Document",font=font, command=spc)
            btn_spc_doc.grid(row=1,column=0)
            btn_source_code = Button(frm_log_main2,text="Source Code",font=font, command=show_soruce_code)
            btn_source_code.grid(row=2, column=0)


        def main_frm():
            nonlocal frm_log_main,str_var   #   to reset the window
            frm_log_main = Frame(login_window, bg="gray")
            frm_log_main.grid(row=0,column=0)

            frm_1 = Frame(frm_log_main)
            frm_1.grid(row=0,column=0)
            lab_log1 = Label(frm_1,text='Password',font=font,width=width_label+1)
            str_var = StringVar()
            ent_log1 = Entry(frm_1,font=font,width=width_entry+2,textvariable=str_var,show="*")

            btn_log = Button(frm_log_main,text="log in",command=login,font=font,width=width_label+width_entry)
            btn_log2 = Button(frm_log_main,text="Reset Password",command=reset_password,font=font,width=width_label+width_entry)
            btn_exit = Button(frm_log_main, text="Exit", font=font, width=(width_label + width_entry ), bg="yellow",
                              command=login_window.destroy)
            lab_log1.grid(row=0,column=0)
            ent_log1.grid(row=0,column=1)
            btn_log.grid(row=1,column=0)
            btn_log2.grid(row=2,column=0)
            btn_exit.grid(row=3,column=0)
            login_window.mainloop()
        main_frm()

    @classmethod
    def cmd(cls):
        # os.path.expanduser("~/Desktop") # similar code for future document
        # pa = os.environ['USERPROFILE']
        homePath = os.path.normpath(os.path.expanduser("~"))
        cng = f"cd {homePath} && {os.path.splitdrive(homePath)[0]}"

        resize_cmd = "mode con: cols=50 lines=10"
        # not used cause it removes the scrollbar cause here lines=10 means buffer = 10

        for ip in extraInfo.get_ip_list():
            commands = ['start', 'cmd.exe', "/k", f'title {ip[0]} &&{cng} && ping {ip[1]} -t']
            # proc = subprocess.Popen(c, stdin=subprocess.PIPE, stdout=subprocess.PIPE, shell=True)
            proc = subprocess.Popen(commands, shell=True)
        return None

    @classmethod
    def cmdSetting(cls):
        # not have any usability, useless function..
        count = 0
        thread_list = []
        def myFun(ip=""):
            pp = subprocess.Popen(f"ping -n 1 {ip}", stdout=subprocess.PIPE, stdin=subprocess.PIPE, shell=True,
                                  universal_newlines=True, stderr=subprocess.PIPE)
            for i in range(1, 2):
                nonlocal count
                time.sleep(1)
                output, error = pp.communicate()
                pp.wait(30)
                firstIP = output.split("\n")[2]
                print("first IP: ", firstIP)
                printCon(f"first IP: , {firstIP}")
                receive = output[output.find('Received = ') + 'Received = '.__len__()]
                receive = int(receive)
                lost = output[output.find('Lost = ') + 'Lost = '.__len__()]
                lost = int(lost)
                print(f"IP address: {ip}; Received: {receive}; Lost: {lost} ")
                printCon(f"IP address: {ip}; Received: {receive}; Lost: {lost} ")
                if lost==1:
                    count+=1

        for i in extraInfo.get_ip_list():
            t1 = threading.Thread(target=myFun, args=(f"{i[1]}",),daemon=True)
            # by using "daemon" keyword, it will kill thread immediately after exit main program
            thread_list.append(t1)
            t1.start()

        t1.join()
        time.sleep(1.5)
        print(count,thread_list)
        printCon(f"{count}:,{thread_list}")
        if count == 0:
            messagebox.showinfo(title="Link Status Summary",message="All links are up..\n(here use only one echo)",parent=top)
        else:
            messagebox.showinfo(title="Link Status Summary",message=f"{count} links are down..\n(here use only one echo)",parent=top)

    print("")

class Options:   # Menu make
    def __init__(self):
        pass

    @classmethod
    def headerColor(cls):
        conn = sqlite3.connect("all_data.db")
        cursor = conn.cursor()
        # cursor.execute("select header_color, side_color from all_info where id = 1")

        global header_color # must use global keyword if want to change it's value.
        # make required folders
        # cwd = os.getcwd()  # reserve current directory
        # if not os.path.exists('data'):
        #     os.mkdir('data')
        # os.chdir("data")
        # if not os.path.exists('color'):
        #     os.mkdir('color')
        # os.chdir(cwd)  # make current directory as previous

        color_temp = askcolor(color="red", title="select Header color")
        if color_temp[0] != None:
            header_color = color_temp[1][1:]
            # file_1 = open(r'data/color/header_color.txt', "w+")
            # file_1.write(color_temp[1][1:])
            cursor.execute(f"update all_info set header_color = '{header_color}' where id = 1;")
            conn.commit()
        return None

    @classmethod
    def sideColor(cls):
        conn = sqlite3.connect("all_data.db")
        cursor = conn.cursor()
        # cursor.execute("select header_color, side_color from all_info where id = 1")

        global side_color
        # # make required folders
        # cwd = os.getcwd()  # reserve current directory
        # if not os.path.exists('data'):
        #     os.mkdir('data')
        # os.chdir("data")
        # if not os.path.exists('color'):
        #     os.mkdir('color')
        # os.chdir(cwd)  # make current directory as previous

        color_temp = askcolor(color="red", title="select side color")
        if color_temp[0] != None:
            side_color = color_temp[1][1:]
            # file_1 = open(r'data/color/side_color.txt', "w")
            # file_1.write(color_temp[1][1:])
            cursor.execute(f"update all_info set side_color = '{side_color}' where id = 1;")
            conn.commit()
        return None

    @classmethod
    def current_d(cls):

        global out_folder_dir
        auto_mail = ''
        print(header_color)
        printCon(header_color)

        if len(os.getcwd()) > len(out_folder_dir):
            abc1 = os.getcwd()
        else:
            abc1=out_folder_dir
        abc = [len("Current Directory: ") ,len(abc1)]
        abc2 = int((abc[0]+abc[1])/2)

        top = Toplevel()

        # width_of_window = 20+len("df")
        # height_of_window = 210
        # x_coordinate = top.winfo_screenwidth() / 2 - width_of_window / 2
        # y_coordinate = top.winfo_screenheight() / 2 - height_of_window / 2
        # top.geometry("+%d+%d" % ( x_coordinate, y_coordinate))

        top.title("Show Settings")
        top.resizable(0, 0)
        top.attributes('-topmost', 'true')  # appear above the main root window

        # add Menu
        menubar = Menu(top)
        noc_name = Menu(menubar, tearoff=False)
        noc_name.add_radiobutton(label="Habib", command=partial(mail_instance.__init__, "Md. Habibur Rahman"))
        noc_name.add_radiobutton(label="Turzo", command=partial(mail_instance.__init__, "Akil Monsur"))
        noc_name.add_radiobutton(label="Turaz", command=partial(mail_instance.__init__, "Tanzil Monsur"))
        noc_name.add_radiobutton(label="Amit", command=partial(mail_instance.__init__, "Amit Roy"))
        noc_name.add_radiobutton(label="Tusher", command=partial(mail_instance.__init__, "Md. Tasnim Rahman Tusher"))
        noc_name.add_radiobutton(label="Obeyddullah",
                                 command=partial(mail_instance.__init__, "A.E.M Obeyddullah Siddique"))
        menubar.add_cascade(menu=noc_name, label="Employee")
        top.config(menu=menubar)

        color_1 = "#c9a836"
        color_2 = "#c7c438"
        fram_top = Frame(top)
        frm = Frame(fram_top, bg="white", bd=5)
        Label(frm, text="Current Directory: ", font=font,bg=color_1, width=abc[0],height=1)\
            .grid(row=2, column=0)
        Label(frm, text=f' {os.getcwd()} ', font=font,bg=color_1, width=abc[1],height=1)\
            .grid(row=2, column=1)

        global text_var_out_folder
        text_var_out_folder.set(out_folder_dir)
        Button(frm, text="Output Folder: ", font=font, bg=color_2, width=abc[0], height=1
               ,command=Options.change_out_folder).grid(row=3, column=0)
        Button(frm, textvariable=text_var_out_folder, font=font,bg=color_2,command=Options.change_out_folder,
               height=1, width=abc[1]).grid(row=3, column=1)
        #################################

        # frm2 = Frame(fram_top, bg="white", bd=5)

        Label(frm, text=f"Employee : ", font=font,bg=color_1,width=abc[0])\
            .grid(row=4, column=0)
        Label(frm, textvariable=text_var_emp_name, font=font,bg=color_1,width=abc[1])\
            .grid(row=4, column=1)

        def update_enable_mail_Var():
            if auto_mail_enable.get() == 1:
                auto_mail_enable.set(0)
                ato = "Disable"
            else:
                auto_mail_enable.set(1)
                ato = "Enable"
            text_var_auto_mail_e.set(ato)

        Button(frm, text=f"Auto Mailing : ",command=update_enable_mail_Var, font=font,bg=color_2,width=abc[0])\
            .grid(row=5, column=0)
        Button(frm, textvariable=text_var_auto_mail_e,command=update_enable_mail_Var, font=font,bg=color_2,width=abc[1])\
            .grid(row=5, column=1)

        def monthly_kpi_summary_control():
            nonlocal monthly_kpi_summary_text
            nonlocal conn,cursor
            cursor.execute("select summary_anable from  all_info where id = 1;")
            monthly_kpi_summary_enable = cursor.fetchone()[0]

            if monthly_kpi_summary_enable == 1:
                conn.execute("UPdate all_info set summary_anable = 0 where id = 1;")
                conn.commit()
                ato = "Disable"
                print("inside disable")
            else:
                print("inside enable")
                conn.execute("UPdate all_info set summary_anable = 1 where id = 1;")
                conn.commit()
                ato = "Enable"

            monthly_kpi_summary_text.set(ato)

        monthly_kpi_summary_text = StringVar(); monthly_kpi_summary_text.set("Disable")
        # global cursor,conn
        conn = sqlite3.connect("all_data.db")
        cursor = conn.cursor()
        cursor.execute("select summary_anable from  all_info where id = 1;")
        enable_summary = cursor.fetchone()[0]

        if enable_summary == 0:monthly_kpi_summary_text.set("Disable")
        else: monthly_kpi_summary_text.set("Enable")

        Button(frm, text=f"Monthly KPI Summary ",command=monthly_kpi_summary_control, font=font,bg=color_2,width=abc[0])\
            .grid(row=6, column=0)
        Button(frm, textvariable=monthly_kpi_summary_text,command=monthly_kpi_summary_control, font=font,bg=color_2,width=abc[1])\
            .grid(row=6, column=1)

        frm.grid(row=0, column=0)
        # frm2.grid(row=1, column=0)

        # This two functions are used only for dynamically change the Setting page color.
        # This function reload the Button with new header or side color.
        def header_func():
            Options.headerColor()
            btn_h.configure(bg = f'#{header_color}')

        def side_func():
            Options.sideColor()
            btn_s.configure(bg=f'#{side_color}')

        frm3 = Frame(fram_top)
        btn_h = Button(frm3, text="Header color",width=abc2, command=header_func,bg=f"#{header_color}", font=font, bd=5)
        btn_h.grid(row=4, column=0)
        btn_s = Button(frm3, text="Side color",width=abc2, command=side_func, bg=f"#{side_color}", font=font, bd=5)
        btn_s.grid(row=4, column=1)
        frm3.grid()

        btn1 = Button(fram_top, text="Exit",bg="#fab005", font=font,bd=bd,
                      width=abc2*2+2,command=partial(lambda: top.destroy()))
        btn1.grid(row=4, column=0)
        fram_top.grid(row=0, column=0)
        return None

    @staticmethod
    def change_out_folder():
        conn = sqlite3.connect("all_data.db")
        cursor = conn.cursor()
        global out_folder_dir,out_folder,text_var_out_folder
        out_folder_dir_1 = filedialog.askdirectory(title = "Select Output Folder")

        if out_folder_dir_1 != "":
            conn.execute(f"update all_info set out_folder_dir  = '{out_folder_dir_1}' where id = 1")
            conn.commit()
            out_folder_dir =out_folder_dir_1
            out_folder = out_folder_dir_1

            # make required folders
            # cwd = os.getcwd()  # reserve current directory
            # if not os.path.exists('data'):
            #     os.mkdir('data')
            # os.chdir("data")
            # if not os.path.exists('output dir'):
            #     os.mkdir('output dir')
            # os.chdir(cwd)  # make current directory as previous
            # file_1 = open(r'data/output dir/outFolder.txt', "w")
            # file_1.write(out_folder_dir_1)
            # file_1.flush()
            # file_1.close()
            # # out_folder=out_folder_dir_1
            text_var_out_folder.set(out_folder_dir_1)
            print('out_folder_dir : ',out_folder_dir_1,type(out_folder_dir))
            printCon(f"out_folder_dir : {out_folder_dir_1},{type(out_folder_dir)}")
        else:pass
        return None

    print()

class extraInfo:
    def __init__(self):
        pass

    @staticmethod
    def formulas():
        pass

    @staticmethod
    def get_ip_list():
        # this function is used by Ping setting class
        ip_list = [
            ["DHA - CTG(SCL)", "172.31.40.22",1,1],
            ["DHA - CTG(SCL)", "172.31.40.21",1,2],

            ["DHA-CTG(F@H)", "172.31.50.10",1,3],
            ["DHA-CTG(F@H)", "172.31.50.9",1,4],

            ["DHA-KHL(F@H)", "10.30.50.10",1,5],
            ["DHA-KHL(F@H)", "10.30.50.9 ",1,6],

            ["DHA-KHL(SCL)", "10.30.40.22",1,7],
            ["DHA-KHL(SCL)", "10.30.40.21",1,8]
        ]
        return ip_list

    @staticmethod
    def show_image_doc():
        image_number=0
        def next():
            nonlocal image_number
            print("next",image_number)
            printCon(f'"next",{image_number}')

            if image_number < len(all_image_names)-1:
                image_number += 1
                set_image(image_number)
            else:
                image_number = 0
                set_image(image_number)

        def previous():
            nonlocal image_number
            print(f"previous,{image_number}")
            printCon(f"previous,{image_number}")
            if image_number > -len(all_image_names)+1:
                image_number -= 1
                set_image(image_number)
            else:
                image_number = len(all_image_names) - 1
                set_image(image_number)

        def set_image(var):
            label_image.configure(image=image_var[var])
            nonlocal root
            root.title("Document")


        def resize(var):
            try:
                image1 = Image.open(var)
                if image1.size[0] != 800:
                    cover = resizeimage.resize_cover(image1, [800, 600])
                    old_path = os.getcwd()
                    os.chdir(os.path.split(var)[0])
                    cover.save(os.path.split(var)[1], image1.format)
                    os.chdir(old_path)
            except Exception as e:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                traceback.print_exception(exc_type, exc_value, exc_traceback)
                exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
                [printCon(i, color='red') for i in exc]
                print(e)
                printCon(f"Exception in image show:Type: {e}", color='red')

        root = Toplevel()
        root.geometry('+%d+%d' % (100, 100))
        root.geometry('%dx%d' % (800, 660))
        frm = Frame(root)
        frm_1 = Frame(frm,bg="#ff0000",width=100,height=100)
        frm_2 = Frame(frm,bg="#ff0000",width=100,height=100)
        label_image = Label(frm_1)
        label_image.grid(row = 0,column=0)

        btn_1 = Button(frm_2,text="Previous",command=previous,width=30,font=font,relief=RAISED,bg='#84dd22')
        btn_2 = Button(frm_2)
        btn_3 = Button(frm_2,text="Next",command=next,width=30,font=font,relief=RAISED,bg='#84dd22')
        btn_1.grid(row = 0,column=0)
        # btn_2.grid(row = 0,column=1)
        btn_3.grid(row = 0,column=3)
        frm.grid(row = 0,column=0)
        frm_1.grid(row = 0,column=0)
        frm_2.grid(row = 1,column=0)

        path = os.path.normpath(os.getcwd() + os.path.sep + "data/image_doc")
        all_image_names = os.listdir(path)
        print(all_image_names)
        printCon(f"all image name: {all_image_names}")
        image_var = []
        # def set_image(i):
        for i in all_image_names: # dont know why it must be done first... it will not done after start software
            temp_path = f"{path}/{i}"
            image1 = Image.open(temp_path)
            resize(temp_path)
            image_var.append(ImageTk.PhotoImage(image1))

        set_image(image_number)

        root.mainloop()
        return None

    @staticmethod
    def about():
        about_top = Toplevel(relief=tkinter.GROOVE)
        about_top.title("About me")
        # about_top.lift(top)
        # about_top.resizable(width=False, height=False) #
        about_top.attributes("-toolwindow", 1)  # It removes both minimize and maximize button
        about_top.attributes('-topmost', 'true')    # appear above the main root window
        width_of_window = 450
        height_of_window = 225
        x_coordinate = top.winfo_screenwidth() / 2 - width_of_window / 2
        y_coordinate = top.winfo_screenheight() / 2 - height_of_window / 2
        about_top.geometry("%dx%d+%d+%d" % (width_of_window, height_of_window, x_coordinate, y_coordinate))
        path1 = os.path.normpath(os.getcwd() + os.path.sep + "data/image_about")
        path = os.path.normpath(path1 + os.path.sep + "habib.jpg")

        if os.path.isdir(path):
            pass
        elif not os.path.isdir(os.path.split(os.path.split(path)[0])[0]):
            os.mkdir("data")
            os.chdir(os.path.normpath(os.getcwd() + os.path.sep + "data"))
            os.mkdir("image_about")
        elif not os.path.isdir(os.path.split(path)[0]):
            os.chdir(os.path.normpath(os.getcwd() + os.path.sep + "data"))
            os.mkdir("image_about")

        if not os.path.isfile(os.path.normpath(os.path.split(path)[0] + os.path.sep + "info.txt")):
            os.chdir(os.path.split(path)[0])
            with open("info.txt", "w") as info_file:
                info_file.write("This folder used to store image named 'habib.jpg', will use in 'about me' Window. ")

        os.chdir(current_working_dir)

        frm   = Frame(about_top,width=width_of_window,height=height_of_window,bg="#ffffff")
        frm_1 = Frame(frm,width=200,height=200)
        frm_2 = Frame(frm)

        text1 = "Md. Habibur Rahman \n M M Communications Ltd. \nMail: md_habibur@outlook.com \n Phone: 01749036120" \
                "\n             01929901565 \n\n Education: \n Electrical and Electronic Engineering \n" \
                " CCNA and CCNP Certified  \nVerification No:\n DNG7WZTHCKV4175H"
        lab = Label(frm_2,bg="#ffffff",font='Times 12',text=text1)
        lab.pack()
        try:
            image1 = Image.open(path)
            # root.geometry('%dx%d' % (image1.size[0], image1.size[1]))
            if image1.size[0]!=200:
                cover = resizeimage.resize_cover(image1, [200, 200])
                old_path = os.getcwd()
                os.chdir(path1)
                cover.save('habib.jpg', image1.format)
                os.chdir(old_path)
            tkpi = ImageTk.PhotoImage(image1)
            label_image = Label(frm_1, image=tkpi)
            label_image.place(x=0, y=0, width=image1.size[0], height=image1.size[1])
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(e)
            printCon(f"inside About Me... Exception type: {e}", color='red')

        frm_1.grid(row=0,column=0)
        frm_2.grid(row=0,column=1)
        frm.pack()
        about_top.mainloop()

    @staticmethod
    def callback():
        if messagebox.askokcancel("Quit", "Do you really wish to quit?"):
            top.destroy()

    print()

class Mail:
    font_size = 14.5

    def __init__(self,name="Md. Habibur Rahman"):
        global text_var_emp_name
        print(auto_mail_enable.get()) # test "auto_mail_enable" value
        printCon(f"atuo mail enable or not indicator: {auto_mail_enable.get()}") # test "auto_mail_enable" value
        printCon(f'Employee name: "{name}" selected..')
        self.name= name
        text_var_emp_name.set(name)
        self.header_sty = '<!DOCTYPE html>' \
                '<html lang="en">'  \
                '<head>'    \
                '<meta charset="UTF-8">'    \
                r'<style> body { font-family: Times New Roman;font-size: ' + f'{Mail.font_size}' + r'px;} </style>' \
                '</head>'   \
                "<body>"
        path_img = os.path.normpath(os.getcwd()+os.sep+ r'data\image\mmcl.jpg')
        self.sign = '<br><br><br>' \
                '<b>Best Regards,</b>  <br>' \
                f'{name}<br>' \
                r'Engineer | Network Operation Center(NOC) <br>' \
                f'<img src="{path_img}">' \
                r'<br>E-mail: noc@mmclbd.com<br>' \
                r'Call us: +8801777189722<br>' \
                '</body>' \
                r'</html>'

    def creat_mail(self, to=None, subject=None, cc=None, body=None, bcc=None, attach=None):
        try:
            outlook = win32.Dispatch('outlook.application')
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            printCon(f"MS OutLook not installed in this machine. {e}", color='red')
            return None
        mail = outlook.CreateItem(0)

        if to != None:
            mail.To = to
        if cc != None:
            mail.cc = cc
        if bcc!= None:
            mail.bcc=bcc
        if subject != None:
            mail.Subject = subject
        if attach != None:
            mail.Attachments.Add(attach)
        if body != None:
            mail.HtmlBody = body
        print("this is create")
        mail.display(False)   # putting True creates mistake
        # mail.send
        return None

    def kpi(self):
        to="raj@mmclbd.com"
        cc="sharma.chandan@mmclbd.com; anirban@mmclbd.com; arif@mmclbd.com; noc@mmclbd.com"
        try:
            attach = os.path.normpath(out_folder_dir + os.sep + mmcl_domestic_ISD_report_class.nam_temp_kpi)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"Error type in KPI : {e}")
            printCon(f"Error type in KPI : {e}", color='red')
            attach = None

        c_time = time.localtime()
        try:
            time_t = mmcl_domestic_ISD_report_class.nam_temp_kpi_t  # [year, month, day, upto time]
        except Exception: time_t = [c_time[0],month_name[c_time[1]],c_time[2],'']

        subject = f"ICX KPI report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}"
        body = self.header_sty + \
            'Dear Vaiya, <br><br>'  \
            f'Please check the attached KPI report of ICX domestic call on {time_t[2]} {time_t[1]} {time_t[0]}{time_t[3]}.' \
            + self.sign

        print(body)
        printCon(f"mail body : {body}")
        # obj_body = open("mail_kpi_m.html","r")
        # body = obj_body.read()
        self.creat_mail(to=to,cc=cc,subject=subject,body=body,attach=attach)
        return None

    def ios_idd(self):
        to = "raj@mmclbd.com"
        cc = "sharma.chandan@mmclbd.com; anirban@mmclbd.com; arif@mmclbd.com; noc@mmclbd.com"
        # set attached file name.
        try:
            attach = os.path.normpath(out_folder_dir + os.sep + mmcl_domestic_ISD_report_class.nam_temp_ios)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("Error type in KPI : ", e)
            printCon(f"Error type in KPI : {e}", color='red')
            attach = None

        # set time and Date
        c_time = time.localtime()
        try:
            time_t = mmcl_domestic_ISD_report_class.nam_temp_ios_t  # [year, month, day, upto time]
        except Exception:
            time_t = [c_time[0], month_name[c_time[1]], c_time[2], '']

        subject = f"IOS & ISD Report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached IDD incoming ( IOS & ISD) report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}.' \
               + self.sign

        print(f"mail body IOS: {body}")
        printCon(f"mail body IOS: {body}")
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, attach=attach)
        return None

    def idd(self):
        to = "raj@mmclbd.com"
        cc = "sharma.chandan@mmclbd.com; anirban@mmclbd.com; arif@mmclbd.com; noc@mmclbd.com"

        # set attached file name.
        try:
            attach = os.path.normpath(out_folder_dir + os.path.sep + mmcl_domestic_ISD_report_class.nam_temp_idd)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print("Error type in KPI : ", e)
            printCon(F"Error type in KPI : {e}", color='red')
            attach = None

        # set time and Date
        c_time = time.localtime()
        try:
            time_t = mmcl_domestic_ISD_report_class.nam_temp_idd_t  # [year, month, day, upto time]
        except Exception:
            time_t = [c_time[0], month_name[c_time[1]], c_time[2], '']

        subject = f"IDD Report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached IDD Report on {time_t[2]} {time_t[1]} {time_t[0]} {time_t[3]}.' \
               + self.sign
        print(body)
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, attach=attach)
        return None

    def idd_every_two_hours(self):
        to = "raj@mmclbd.com"
        cc = "anirban@mmclbd.com; arif@mmclbd.com; noc@mmclbd.com"

        # set attached file name.
        try:
            attach = out_folder_dir + os.sep + idd_class.filename_mail_2h
            attach = os.path.normpath(attach)
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"Error type in KPI : {e}")
            printCon(f"Error type in KPI : {e}", color='red')
            attach = None

        # set time and Date
        c_time = time.localtime()
        try:
            time_t = idd_class.filename_mail_2h_t  # [year, month, day, upto_time]
        except Exception:
            time_t = [c_time[0], month_name[c_time[1]], c_time[2], f'{time.strftime("%I",c_time)} {time.strftime("%p",c_time)}']
            # Last element of this list will print like "10 AM"

        subject = f"IDD Report on {time_t[2]} {time_t[1]} {time_t[0]} up to {time_t[3]}"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached IDD Report on  {time_t[2]} {time_t[1]} {time_t[0]} up to {time_t[3]}.' \
               + self.sign
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, attach=attach)
        return None

    def Peak_Hour_Traffic_Statistics(self):
        to = "raj@mmclbd.com"
        cc = "anirban@mmclbd.com; arif@mmclbd.com; noc@mmclbd.com"

        # set time and Date
        c_time = time.localtime()

        time_t = [c_time[0], month_name[c_time[1]], c_time[2],
                  f'{time.strftime("%I", c_time)} {time.strftime("%p", c_time)}']
        # Last element of this list will print like "10 AM"

        subject = f"Peak Hour Traffic Statistics Report on {time_t[2]} {time_t[1]} {time_t[0]} (8.30pm to 9.30pm)"
        body = self.header_sty + \
               'Dear Vaiya, <br><br>' \
                f'Please check the attached Peak Hour Traffic Statistics report on {time_t[2]} {time_t[1]} {time_t[0]}.' \
               + self.sign
        print(body)
        printCon(f"Peak hour statistics mail body: {body}")
        self.creat_mail(to=to, cc=cc, subject=subject, body=body)
        return None

    def all_ios(self):
        to = "raj@mmclbd"
        cc = "anirban@mmclbd;arif@mmclbd;noc@mmclbd"
        bcc = ""
        subject = f"IDD Report on "
        body = self.header_sty + \
                self.sign
        self.creat_mail(to=to, cc=cc, subject=subject, body=body, bcc=bcc)
        return None

    print()

class MyThread(threading.Thread):

    def __init__(self,function):
        threading.Thread.__init__(self)
        self.function = function
        self.daemon=True

    def run(self):
        import traceback
        print(f"{threading.current_thread()} started.....................")
        printCon(f"{threading.current_thread()} started.....................")
        threadLock_ins.acquire()
        # remove this try: block if need any analysis. otherwise no exception will show  . . . .
        try:
            self.function()
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback)
            exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
            [printCon(i, color='red') for i in exc]
            print(f"New Trunk may be added today... or Excel file may not closed.. EX type: {e}")
            printCon(f"New Trunk may be added today... or Excel file may not closed.. EX : {e}")
        finally:
            threadLock_ins.release()
        print(f"{threading.current_thread()} ended.......................")
        printCon(f"{threading.current_thread()} ended.......................")

    print("")

class AbstractClass:

    @classmethod
    def every_hour_idd(self):
        # if already this function in active then stop new request
        if "every_hour_idd" in [i.getName() for i in threading.enumerate()]:
            printCon(f"Every hour IDD(International Direct Dialing) report already in queue. wait for a while..", color='red')
            return None

        thread_mmcl = MyThread(instance_idd_class.csv_file)
        thread_mmcl.setName("every_hour_idd")
        thread_mmcl.start()

    @classmethod
    def ccr_check(self):
        if "ccr_check" in [i.getName() for i in threading.enumerate()]:
            printCon(f"CCR (Call Completion Rate) check already in queue. wait for a while..(ZTE: Connect No / Attempt No)", color='red')
            return None
        thread_mmcl = MyThread(instance_ccr_class.csv_file)
        thread_mmcl.setName("ccr_check")
        thread_mmcl.start()

    @classmethod
    def in_comming_kpi(self):
        if "in_comming_kpi" in [i.getName() for i in threading.enumerate()]:
            printCon(f"In comming KPI(Key Performance Indicator) report already in queue. wait for a while..", color='red')
            return None
        thread_mmcl = MyThread(raw_kpi_in_instance.make_kpi_report)
        thread_mmcl.setName("in_comming_kpi")
        thread_mmcl.start()

    @classmethod
    def out_going_kpi(self):
        if "out_going_kpi" in [i.getName() for i in threading.enumerate()]:
            printCon(f"Out going KPI(Key Performance Indicator) report already in queue. wait for a while..", color='red')
            return None
        thread_mmcl = MyThread(raw_kpi_out_instance.make_IOS_ISD_report)
        thread_mmcl.setName("out_going_kpi")
        thread_mmcl.start()

    @classmethod
    def idd_day_calculation_time_period(self):
        if "idd_day_calculation_time_period" in [i.getName() for i in threading.enumerate()]:
            printCon(f"IDD Day of Calculation Time Period report already in queue. wait for a while..", color='red')
            return None
        thread_mmcl = MyThread(raw_idd.make_IDD_report)
        thread_mmcl.setName("idd_day_calculation_time_period")
        thread_mmcl.start()

    print("hello,, abstract class")


def setting_var():
    # this two StringVar used to Show current Settings.
    global text_var_emp_name, text_var_auto_mail_e, text_var_out_folder,instance_idd_class,instance_ccr_class
    text_var_emp_name = StringVar()
    text_var_auto_mail_e = StringVar()
    text_var_auto_mail_e.set("Disable")
    text_var_out_folder = StringVar()
    instance_idd_class = idd_class();
    instance_ccr_class = ccr_class();
    if True:
        c1 = sqlite3.connect("all_data.db")
        c1.execute("create table if not exists all_info "
                   "(id INT PRIMARY KEY NOT NULL,"
                   "header_color text not null, "
                   "side_color text not null,"
                   "out_folder_dir text,"
                   "password text,"
                   "summary_anable int);")
        c1.execute("INSERT OR IGNORE INTO all_info (id, header_color, side_color, summary_anable,password)"
                   "values (1, 'ff0000', '00ff00' , 0, 'password');")
        c1.commit()
        c1.close()

    global thread_list,threadLock_ins
    thread_list = ["every_hour_idd","ccr_check","in_comming_kpi","out_going_kpi","idd_day_calculation_time_period"]
    threadLock_ins = threading.Lock()
    global ip_list
    ip_list = [
        ["DHA - CTG(SCL)", "172.31.40.22"],
        ["DHA - CTG(SCL)", "172.31.40.21"],

        ["DHA-CTG(F@H)", "172.31.50.10"],
        ["DHA-CTG(F@H)", "172.31.50.9"],

        ["DHA-KHL(F@H)", "10.30.50.10"],
        ["DHA-KHL(F@H)", "10.30.50.9 "],

        ["DHA - KHL(SCL)", "10.30.40.22"],
        ["DHA - KHL(SCL)", "10.30.40.21"]
    ]
    print(f"bok .. . {ip_list}")
    return None

if __name__ == "__main__":

    top = Tk()
    # this function must call before Open_gui
    setting_var()

    conn = sqlite3.connect("all_data.db")
    cursor = conn.cursor()
    cursor.execute("select header_color, side_color from all_info where id = 1")
    color_db = cursor.fetchone()
    print(color_db," colors . . . . . . ... . /// / / / ")
    header_color = color_db[0] # header color of excel file
    side_color = color_db[1]   # side/result color of excel file

    # Store the Current directory of software
    current_working_dir = os.getcwd()
    try:
        cursor.execute("select out_folder_dir from all_info where id = 1")
        # out_folder = os.path.normpath(open(r'data/output dir/outFolder.txt', "r").read())
        out_folder = cursor.fetchone()[0]
        if os.path.isdir(out_folder):
            out_folder_dir = out_folder
        else: out_folder_dir = current_working_dir
    except Exception:
        out_folder = "not a dir"
        out_folder_dir = current_working_dir


    top.protocol("WM_DELETE_WINDOW", extraInfo.callback)
    top.title("MMCL Periodic Monitoring");
    try:
        path_1 = os.path.normpath(os.getcwd() + os.sep + r'data\image\mmcl.ico')
        # print(path_1)
        if os.path.exists(path_1):
            top.iconbitmap(default=path_1)
        elif os.path.exists(r'C:\mmcl.ico'):
            top.iconbitmap(default = r'C:\mmcl.ico');
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        traceback.print_exception(exc_type, exc_value, exc_traceback)
        exc = traceback.format_exception(exc_type, exc_value, exc_traceback)
        [printCon(i, color='red') for i in exc]
        print(f"App logo not found !!. Exception type: {e}")
        # printCon(f"App logo not found !!. Exception type: {e}")


    top.resizable(width=False,height=False);

    width_of_window = 832;
    height_of_window = 377;
    screen_width = top.winfo_screenwidth();
    screen_height = top.winfo_screenheight();
    x_coordinate = screen_width/2 - width_of_window/2;
    y_coordinate = screen_height/2 - height_of_window/2;
    top.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))

    top.configure(bg="#ffffff")  #set app background
    bg_1 = "#a2d32c"
    bg_11 = "#afc43c"
    bg_2 = "#7bdc23"
    bg_21 = "#7bdc2f"
    fg_1 = "#000000"
    bd=3
    relief = "raised"
    width_1 = 70
    font = "Times 15"


    # Menu make
    Open_gui.menu_bars()

    frm = Frame(top,bd=5,relief="solid",pady=5,padx=5)
    frm_1 = Frame(frm,bd=5,relief="solid",width=500)
    frm_2 = Frame(frm,bd=5,relief="solid",width=500)

    var_1 = StringVar(); var_1.set(instance_idd_class.filename);
    var_3 = StringVar(); var_3.set("Make IDD Report");

    btn_1 = Button(frm_1, textvariable=var_1, command=instance_idd_class.upload, width=width_1, font=font,
                   bg=bg_1, fg=fg_1, relief=relief, bd=bd)
    btn_3 = Button(frm_1, textvariable=var_3, command=AbstractClass.every_hour_idd, width=width_1, font=font,
                   bg=bg_1, relief=relief, bd=bd, fg=fg_1);
    Label(frm_1,text="IDD Every Two hours",bg=bg_11,width=width_1+1,font=font,height=2).grid(row=0,column=0)
    btn_1.grid(row=1,column=0)
    btn_3.grid(row=3,column=0)
    frm_1.grid(row=1,column=0)       # row=1 in main frame "frm"

    Label(frm,width=115).grid(row=0,column=0) # up side blank space      # row=0 in main frame "frm"
    Label(frm,width=115).grid(row=2,column=0) # buttom side blank space  # row=2 in main frame "frm"

    var1 = StringVar(); var1.set(instance_ccr_class.filename);
    var2 = StringVar(); var2.set("CCR Calculate");
    btn1 = Button(frm_2, textvariable=var1, command=instance_ccr_class.upload, width=width_1, font=font,
                  relief=relief, bd=bd, bg=bg_2, fg=fg_1);
    btn2 = Button(frm_2, textvariable=var2, command=AbstractClass.ccr_check, width=width_1, font=font,
                  relief=relief, bd=bd, bg=bg_2, fg=fg_1);
    # CCR Check Header
    frm_check_b = Frame(frm_2,bg=bg_21)
    Label(frm_check_b,text=" CCR Check ",bg=bg_21,width=width_1//2+6,font=font,height=2, fg=fg_1).grid(row=0,column=0)
    Button(frm_check_b, text="Select Countries", command=instance_ccr_class.country_selection, font=font, height=1, fg=fg_1, bg=bg_21).grid(row=0, column=1)
    intVar_ccr_auto_open = IntVar();
    Checkbutton(frm_check_b,text="Open in Excel",bg=bg_21,width=width_1//2-22,font=font,height=1, fg=fg_1,onvalue=1,
                offvalue=0,variable=intVar_ccr_auto_open).grid(row=0,column=2)
    frm_check_b.grid(row=0,column=0)

    btn1.grid(row=1,column=0)
    btn2.grid(row=2,column=0)
    frm_2.grid(row=3,column=0)       # row=3 in main frame "frm"

    text_var_out_folder.set(out_folder_dir)
    lab_1 = Label(frm, font=font,bg="white",width=width_1,textvariable=text_var_out_folder)

    lab_1.grid(row=4,column=0)
    frm.grid(row=2,column = 0)      # this is the main frame for ccr idd

    top.mainloop();


